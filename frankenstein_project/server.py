"""
Merged LabelKit backend
-----------------------
This FastAPI app merges two related workflows without mixing their PDF/XML logic.

BACKEND MAP
- /generate/michaels -> Michaels XML + shipping-label PDF workflow.
- /generate/kehe     -> KeHE XML-only GS1 label workflow.
- /results/{id}/*    -> shared async status/report/file endpoints used by both workflows.

The frontend lives in frontend/dist/index.html and calls the kit-specific backend
endpoint selected by the user.
"""

from __future__ import annotations

import base64
import json
import csv
import io
import math
import os
import re
import shutil
import tempfile
import threading
import time
import uuid
import urllib.error
import urllib.parse
import urllib.request
import zlib
from pathlib import Path
from typing import Any, Dict, List, Optional

import fitz
from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles

# ---------------------------------------------------------------------------
# BACKEND SECTION 1: import the two separate processing engines.
# Do not merge these modules. Michaels and KeHE parse/render labels differently.
# ---------------------------------------------------------------------------
from pipelines.michaels_label_pipeline import (  # noqa: E402
    MatchFailureError as MichaelsMatchFailureError,
    run_pipeline as run_michaels_pipeline,
)
from pipelines.kehe_pipeline import (  # noqa: E402
    run_pipeline as run_kehe_pipeline,
    build_kehe_master_packing_list_draft,
    build_kehe_pallet_label_draft,
    build_kehe_pack_label_draft,
    render_kehe_master_packing_list_pdf,
    render_kehe_pallet_label_pdf,
    render_kehe_pack_label_pdf,
    load_kehe_dc_directory,
)

MatchFailureErrors = (MichaelsMatchFailureError,)

# ---------------------------------------------------------------------------
# BACKEND SECTION 2: shared app configuration and static frontend hosting.
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).resolve().parent
FRONTEND_DIST = BASE_DIR / "frontend" / "dist"
FRONTEND_ASSETS = FRONTEND_DIST / "assets"
DEFAULT_PORT = int(os.getenv("X_ZOHO_CATALYST_LISTEN_PORT", os.getenv("PORT", "9000")))
APP_NAME = "Merged LabelKit"
APP_ID = "merged-labelkit"
MAX_CACHED_REPORTS = 25

RESULT_REPORTS: Dict[str, Dict[str, Any]] = {}
RESULT_JOBS: Dict[str, Dict[str, Any]] = {}


def _env_bool(name: str, default: bool = False) -> bool:
    raw = os.getenv(name)
    if raw is None:
        return default
    return str(raw).strip().lower() in {"1", "true", "yes", "on"}


APP_CONFIG_FILE = Path(os.getenv("LABELKIT_CONFIG_FILE", str(BASE_DIR / "labelkit_config.json")))


def _load_labelkit_config() -> Dict[str, Any]:
    try:
        if not APP_CONFIG_FILE.exists():
            return {}
        data = json.loads(APP_CONFIG_FILE.read_text(encoding="utf-8"))
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def _is_catalyst_runtime() -> bool:
    catalyst_markers = {
        "X_ZOHO_CATALYST_LISTEN_PORT",
        "X_ZOHO_CATALYST_PROJECT_ID",
        "X_ZOHO_CATALYST_ORG_ID",
        "X_ZOHO_CATALYST_APP_NAME",
        "CATALYST_PROJECT_ID",
        "CATALYST_APP_NAME",
        "CATALYST_OPTIONS",
    }
    return any(os.getenv(marker) for marker in catalyst_markers)


LABELKIT_CONFIG = _load_labelkit_config()
HAS_LABELKIT_CONFIG = bool(LABELKIT_CONFIG)


def _resolve_labelkit_profile() -> tuple[str, Dict[str, Any]]:
    profiles = LABELKIT_CONFIG.get("profiles")
    if not isinstance(profiles, dict):
        profiles = {}
    requested = str(
        os.getenv("LABELKIT_PROFILE")
        or LABELKIT_CONFIG.get("active_profile")
        or "local"
    ).strip().lower()
    if requested == "auto":
        requested = "catalyst" if _is_catalyst_runtime() else "local"
    profile = profiles.get(requested, {})
    return requested, profile if isinstance(profile, dict) else {}


LABELKIT_CONFIG_PROFILE, ACTIVE_LABELKIT_CONFIG = _resolve_labelkit_profile()
ALLOW_CONFIG_ENV_OVERRIDES = bool(
    LABELKIT_CONFIG.get("allow_environment_overrides", not HAS_LABELKIT_CONFIG)
)


def _config_value(env_name: str, key: str, default: Any = "") -> Any:
    if ALLOW_CONFIG_ENV_OVERRIDES:
        raw = os.getenv(env_name)
        if raw is not None:
            return raw
    if key in ACTIVE_LABELKIT_CONFIG:
        return ACTIVE_LABELKIT_CONFIG.get(key)
    return default


def _config_bool(env_name: str, key: str, default: bool = False) -> bool:
    if ALLOW_CONFIG_ENV_OVERRIDES and os.getenv(env_name) is not None:
        return _env_bool(env_name, default)
    raw = ACTIVE_LABELKIT_CONFIG.get(key, default)
    if isinstance(raw, bool):
        return raw
    return str(raw).strip().lower() in {"1", "true", "yes", "on"}


def _config_path(env_name: str, key: str, default: str) -> Path:
    raw = _config_value(env_name, key, default)
    path = Path(str(raw or default))
    return path if path.is_absolute() else BASE_DIR / path


APP_ENV = str(_config_value("APP_ENV", "app_env", os.getenv("ENVIRONMENT", "local"))).strip().lower()
AUTH_REQUIRED = _config_bool("AUTH_REQUIRED", "auth_required", APP_ENV in {"production", "prod"})
AUTH_MODE = str(_config_value("AUTH_MODE", "auth_mode", "embedded" if AUTH_REQUIRED else "none") or "none").strip().lower()
ALLOW_LOCAL_JSON_FALLBACK = _config_bool("ALLOW_LOCAL_JSON_FALLBACK", "allow_local_json_fallback", not AUTH_REQUIRED)
ALLOW_BROWSER_LOCAL_CACHE = _config_bool("ALLOW_BROWSER_LOCAL_CACHE", "allow_browser_local_cache", not AUTH_REQUIRED)

# Zoho Analytics order lookup used by the standalone Packing List & Ti-Hi
# workspace. OAuth tokens are resolved at request time through the Catalyst
# Connection; no Analytics credentials are stored in this repository.
ANALYTICS_ORDER_SOURCE = str(
    _config_value("ANALYTICS_ORDER_SOURCE", "analytics_order_source", "zoho_analytics")
    or "zoho_analytics"
).strip().lower()
ANALYTICS_LOCAL_SOURCE_VALUES = {"file", "csv", "excel"}
_analytics_local_file_value = str(
    _config_value("ANALYTICS_LOCAL_FILE", "analytics_local_file", "") or ""
).strip()
ANALYTICS_LOCAL_FILE: Optional[Path] = None
if _analytics_local_file_value:
    _analytics_local_path = Path(_analytics_local_file_value)
    ANALYTICS_LOCAL_FILE = (
        _analytics_local_path
        if _analytics_local_path.is_absolute()
        else BASE_DIR / _analytics_local_path
    )
ANALYTICS_CONNECTION_LINK_NAME = str(
    _config_value("ANALYTICS_CONNECTION_LINK_NAME", "analytics_connection_link_name", "orderdata")
    or "orderdata"
).strip()
ANALYTICS_API_BASE = str(
    _config_value("ANALYTICS_API_BASE", "analytics_api_base", "https://analyticsapi.zoho.com")
    or "https://analyticsapi.zoho.com"
).strip().rstrip("/")
ANALYTICS_ORG_ID = str(_config_value("ANALYTICS_ORG_ID", "analytics_org_id", "") or "").strip()
ANALYTICS_WORKSPACE_ID = str(
    _config_value("ANALYTICS_WORKSPACE_ID", "analytics_workspace_id", "1436788000013504925")
    or "1436788000013504925"
).strip()
ANALYTICS_VIEW_ID = str(
    _config_value("ANALYTICS_VIEW_ID", "analytics_view_id", "1436788000014668542")
    or "1436788000014668542"
).strip()
ANALYTICS_VIEW_NAME = str(
    _config_value("ANALYTICS_VIEW_NAME", "analytics_view_name", "Data with Product Details")
    or "Data with Product Details"
).strip()
ANALYTICS_ORDER_COLUMN = str(
    _config_value("ANALYTICS_ORDER_COLUMN", "analytics_order_column", "Sales Order Number")
    or "Sales Order Number"
).strip()
ANALYTICS_SKU_COLUMN = str(
    _config_value("ANALYTICS_SKU_COLUMN", "analytics_sku_column", "SKUNumber")
    or "SKUNumber"
).strip()
ANALYTICS_QUANTITY_COLUMN = str(
    _config_value("ANALYTICS_QUANTITY_COLUMN", "analytics_quantity_column", "Quantity Ordered")
    or "Quantity Ordered"
).strip()
ANALYTICS_ORDER_INSTANCE_ID_COLUMN = "Ecomdash ID"
ANALYTICS_ORDER_INSTANCE_DATE_COLUMN = "Invoice Date"
ANALYTICS_ORDER_INSTANCE_STOREFRONT_COLUMN = "Storefront"
ANALYTICS_ORDER_DETAIL_COLUMNS: Dict[str, str] = {
    "billing_customer_name": "Billing Customer Name",
    "bill_to_phone": "Bill To Phone",
    "billing_street1": "Billing Street1",
    "billing_street2": "Billing Street2",
    "billing_street3": "Billing Street3",
    "billing_city": "Billing City",
    "billing_state": "Billing State",
    "billing_zip_code": "Billing Zip Code",
    "billing_country": "Billing Country",
    "ship_to_name": "Ship To Name",
    "ship_to_phone": "Ship To Phone",
    "shipping_street1": "Shipping Street1",
    "shipping_street2": "Shipping Street2",
    "shipping_street3": "Shipping Street3",
    "shipping_city": "Shipping City",
    "shipping_state": "Shipping State",
    "shipping_zip_code": "Shipping Zip Code",
    "shipping_country": "Shipping Country",
    "order_notes": "Order Notes",
}
ANALYTICS_HTTP_TIMEOUT_SECONDS = 30
ANALYTICS_DISCOVERED_ORG_ID = ""

KIT_CONFIG: Dict[str, Dict[str, str]] = {
    "michaels": {
        "label": "Michaels Label Kit",
        "output_filename": "michaels_rollo_output.pdf",
        "temp_prefix": "michaels_labelkit_",
    },
    "kehe": {
        "label": "KeHE Label Kit",
        "output_filename": "kehe_gs1_labels.pdf",
        "temp_prefix": "kehe_labelkit_",
    },
    "kehe_pallet_label": {
        "label": "KeHE Pallet Label",
        "output_filename": "kehe_pallet_labels.pdf",
        "temp_prefix": "kehe_pallet_label_",
    },
    "kehe_master_packing_list": {
        "label": "KeHE Master Packing List",
        "output_filename": "kehe_master_packing_list.pdf",
        "temp_prefix": "kehe_master_packing_list_",
    },
    "kehe_pack_labels": {
        "label": "KeHE Pack Labels",
        "output_filename": "kehe_pack_labels.pdf",
        "temp_prefix": "kehe_pack_labels_",
    },
}

# Product Master persistence.
# MPL_PRODUCT_MASTER_TABLE is the shared source of truth. KeHE reads and writes
# only rows where Storefront = KeHE, preserving other storefront rows.
MPL_PRODUCT_MASTER_TABLE = str(
    _config_value("MPL_PRODUCT_MASTER_TABLE", "mpl_product_master_table", "mpl_product_master")
    or "mpl_product_master"
)
MPL_PRODUCT_MASTER_STORE = str(_config_value(
    "MPL_PRODUCT_MASTER_STORE",
    "mpl_product_master_store",
    "datastore" if AUTH_REQUIRED and not ALLOW_LOCAL_JSON_FALLBACK else "auto",
)).strip().lower()
MPL_PRODUCT_MASTER_FILE = Path(
    os.getenv("MPL_PRODUCT_MASTER_FILE", str(BASE_DIR / "data" / "mpl_product_master.json"))
)

# Package quantities must be explicit in Product Master. Older rows remain
# readable, but Analytics order conversion will not guess a missing case pack.
DEFAULT_CASE_QTY_BY_LEVEL = {
    "Case": "",
    "Inner Pack": "",
    "Each": "",
    "Shipper Contents": "",
    "Other": "",
}
DEFAULT_LABELS_PER_UNIT_BY_LEVEL = {
    "Case": "2",
    "Inner Pack": "6",
    "Each": "",
    "Shipper Contents": "",
    "Other": "",
}
# Default Ship From used only for manual DC Directory rows when no value exists.
DEFAULT_KEHE_SHIP_FROM = "BAKELL LLC\n1967 ESSEX CT\nREDLANDS, CA 92373\nUSA"

# Directory persistence.
# MPL_DIRECTORY_TABLE is the shared source of truth. KeHE uses rows where
# Storefront = KeHE.
MPL_DIRECTORY_TABLE = str(
    _config_value("MPL_DIRECTORY_TABLE", "mpl_directory_table", "mpl_directory")
    or "mpl_directory"
)
MPL_DIRECTORY_STORE = str(
    _config_value("MPL_DIRECTORY_STORE", "mpl_directory_store", MPL_PRODUCT_MASTER_STORE)
).strip().lower()
MPL_DIRECTORY_FILE = Path(
    os.getenv("MPL_DIRECTORY_FILE", str(BASE_DIR / "data" / "mpl_directory.json"))
)
MPL_DRAFTS_TABLE = str(
    _config_value("MPL_DRAFTS_TABLE", "mpl_drafts_table", "kehe_mpl_drafts")
    or "kehe_mpl_drafts"
)
MPL_DRAFTS_STORE = str(_config_value(
    "MPL_DRAFTS_STORE",
    "mpl_drafts_store",
    os.getenv("KEHE_MPL_DRAFTS_STORE", MPL_PRODUCT_MASTER_STORE) if ALLOW_CONFIG_ENV_OVERRIDES else MPL_PRODUCT_MASTER_STORE,
)).strip().lower()
MPL_DRAFTS_FILE = _config_path(
    "MPL_DRAFTS_FILE",
    "mpl_drafts_file",
    os.getenv("KEHE_MPL_DRAFTS_FILE", str(BASE_DIR / "data" / "kehe_mpl_drafts.json")),
)
AUDIT_LOG_TABLE = str(
    _config_value("AUDIT_LOG_TABLE", "audit_log_table", "kehe_audit_log")
    or "kehe_audit_log"
)
AUDIT_LOG_STORE = str(_config_value(
    "AUDIT_LOG_STORE",
    "audit_log_store",
    os.getenv("KEHE_AUDIT_LOG_STORE", MPL_PRODUCT_MASTER_STORE) if ALLOW_CONFIG_ENV_OVERRIDES else MPL_PRODUCT_MASTER_STORE,
)).strip().lower()
AUDIT_LOG_FILE = Path(
    os.getenv(
        "AUDIT_LOG_FILE",
        os.getenv("KEHE_AUDIT_LOG_FILE", str(BASE_DIR / "data" / "kehe_audit_log.json")),
    )
)


app = FastAPI(title=f"{APP_NAME} API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

if FRONTEND_ASSETS.exists():
    app.mount("/assets", StaticFiles(directory=str(FRONTEND_ASSETS)), name="assets")


@app.get("/")
async def root() -> HTMLResponse:
    return serve_frontend_index()


@app.get("/health")
def health() -> Dict[str, Any]:
    return {
        "status": "ok",
        "app": APP_NAME,
        "app_id": APP_ID,
        "frontend_found": (FRONTEND_DIST / "index.html").exists(),
        "frontend_entry": "frontend/dist/index.html",
        "backend_entry": "server.py",
        "available_kits": {
            "michaels": "XML + shipping-label PDF matching workflow",
            "kehe": "XML-only GS1 label workflow",
        },
    }


# ---------------------------------------------------------------------------
# BACKEND SECTION 3: shared job cache used by both kit workflows.
# ---------------------------------------------------------------------------
def _prune_old_results() -> None:
    while len(RESULT_JOBS) > MAX_CACHED_REPORTS:
        oldest_key = next(iter(RESULT_JOBS))
        job = RESULT_JOBS.pop(oldest_key, None)
        RESULT_REPORTS.pop(oldest_key, None)
        if job and job.get("temp_dir"):
            shutil.rmtree(job["temp_dir"], ignore_errors=True)


def create_result_job(temp_dir: Path, kit: str, output_filename: Optional[str] = None) -> str:
    result_id = uuid.uuid4().hex
    RESULT_JOBS[result_id] = {
        "kit": kit,
        "status": "processing",
        "detail": "Files uploaded. Starting generation…",
        "report": None,
        "output_path": None,
        "output_filename": output_filename,
        "temp_dir": str(temp_dir),
        "created_at": time.time(),
    }
    _prune_old_results()
    return result_id


def update_result_job(result_id: str, **changes: Any) -> None:
    job = RESULT_JOBS.get(result_id)
    if job is not None:
        job.update(changes)


# ---------------------------------------------------------------------------
# BACKEND SECTION 4A: Michaels worker.
# Requires XML files and shipping-label PDF files. It uses OCR/matching.
# ---------------------------------------------------------------------------
def run_michaels_generation_job(result_id: str, xml_paths: List[str], pdf_paths: List[str]) -> None:
    job = RESULT_JOBS[result_id]
    temp_dir = Path(job["temp_dir"])
    output_path = temp_dir / KIT_CONFIG["michaels"]["output_filename"]

    try:
        update_result_job(result_id, detail="Combining Michaels shipping-label PDFs…")
        shipping_pdf_path = combine_shipping_pdfs(
            [Path(p) for p in pdf_paths],
            temp_dir / "combined_shipping_labels.pdf",
        )

        def _progress(message: str) -> None:
            update_result_job(result_id, detail=message)

        update_result_job(result_id, detail="Matching Michaels shipping-label pages to XML packs…")
        report = run_michaels_pipeline(
            xml_paths=xml_paths,
            out_pdf=str(output_path),
            shipping_pdf_path=str(shipping_pdf_path),
            progress_callback=_progress,
        )

        if not output_path.exists():
            raise RuntimeError("Output PDF was not generated.")

        RESULT_REPORTS[result_id] = report
        update_result_job(
            result_id,
            status="complete",
            detail="Michaels labels generated successfully.",
            report=report,
            output_path=str(output_path),
        )
    except MatchFailureErrors as exc:
        RESULT_REPORTS[result_id] = exc.report or {}
        update_result_job(result_id, status="error", detail=str(exc), report=exc.report)
    except ValueError as exc:
        update_result_job(result_id, status="error", detail=str(exc))
    except Exception as exc:
        update_result_job(result_id, status="error", detail=str(exc))


# ---------------------------------------------------------------------------
# BACKEND SECTION 4B: KeHE worker.
# Requires XML files only. Shipping-label PDFs are intentionally not used.
# ---------------------------------------------------------------------------
def run_kehe_generation_job(result_id: str, xml_paths: List[str]) -> None:
    job = RESULT_JOBS[result_id]
    temp_dir = Path(job["temp_dir"])
    output_path = temp_dir / KIT_CONFIG["kehe"]["output_filename"]

    try:
        def _progress(message: str) -> None:
            update_result_job(result_id, detail=message)

        update_result_job(result_id, detail="Generating KeHE GS1 labels from XML…")
        report = run_kehe_pipeline(
            xml_paths=xml_paths,
            out_pdf=str(output_path),
            progress_callback=_progress,
        )

        if not output_path.exists():
            raise RuntimeError("Output PDF was not generated.")

        RESULT_REPORTS[result_id] = report
        update_result_job(
            result_id,
            status="complete",
            detail="KeHE labels generated successfully.",
            report=report,
            output_path=str(output_path),
        )
    except MatchFailureErrors as exc:
        RESULT_REPORTS[result_id] = exc.report or {}
        update_result_job(result_id, status="error", detail=str(exc), report=exc.report)
    except ValueError as exc:
        update_result_job(result_id, status="error", detail=str(exc))
    except Exception as exc:
        update_result_job(result_id, status="error", detail=str(exc))



# ---------------------------------------------------------------------------
# BACKEND SECTION 4C: KeHE product master persistence.
# Frontend uses these APIs as the primary source for the GTIN / Packaging table.
# localStorage remains only a browser-side fallback cache.
# ---------------------------------------------------------------------------
def normalize_packaging_level(value: Any) -> str:
    raw = re.sub(r"\s+", " ", str(value or "").strip().lower())
    compact = raw.replace(" ", "")
    if raw in {"case", "cases", "master pack", "master packs", "mp", "case pack", "case packs"} or compact == "casepack":
        return "Case"
    if raw in {"inner pack", "inner packs", "inner", "ip"} or compact in {"innerpack", "innerpacks"}:
        return "Inner Pack"
    if raw in {"each", "ea"}:
        return "Each"
    if raw in {"shipper contents", "shipper content", "shipper", "display shipper"}:
        return "Shipper Contents"
    if raw == "other":
        return "Shipper Contents"
    return "Other"


def _first_value(row: Dict[str, Any], *keys: str) -> str:
    for key in keys:
        if key in row and row.get(key) is not None:
            return str(row.get(key) or "").strip()
    return ""


def _boolish(value: Any, default: bool = False) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return default
    raw = str(value).strip().lower()
    if not raw:
        return default
    if raw in {"1", "true", "yes", "y", "on", "checked", "✅", "x"}:
        return True
    if raw in {"0", "false", "no", "n", "off", "unchecked", "barcode on product"}:
        return False
    if "barcode" in raw and "product" in raw:
        return False
    return default


def _product_in_packing_list(row: Dict[str, Any], packaging_level: str, label_required: str = "") -> bool:
    if normalize_packaging_level(packaging_level) != "Case":
        return False
    for key in ("in_packing_list", "IN_PACKING_LIST", "In Packing List"):
        if key in row and row.get(key) is not None:
            return _boolish(row.get(key), True)
    if label_required:
        return _boolish(label_required, True)
    return True


def _normalize_storefront(value: Any) -> str:
    clean = str(value or "").strip()
    return clean or "KeHE"


def _is_kehe_storefront(value: Any) -> bool:
    return _normalize_storefront(value).lower() == "kehe"


def normalize_product_master_row(row: Dict[str, Any]) -> Dict[str, str]:
    storefront = _normalize_storefront(_first_value(row, "storefront", "STOREFRONT", "Storefront"))
    gtin = _first_value(row, "gtin", "GTIN", "case_upc", "CASE_UPC", "upc", "UPC")
    packaging_level = normalize_packaging_level(
        _first_value(row, "packaging_level", "packging_level", "PACKAGING_LEVEL", "PACKGING LEVEL", "Packaging Level")
    )
    label_required = _first_value(row, "label_required", "LABEL_REQUIRED", "Label Required")
    in_packing_list = _product_in_packing_list(row, packaging_level, label_required)
    case_qty = _first_value(
        row,
        "case_qty",
        "CASE_QTY",
        "Case Qty",
        "Eaches / Package",
        "eaches_per_package",
        "eaches_per_case",
        "eaches_per_inner_pack",
    )
    labels_per_unit = _first_value(row, "labels_per_unit", "LABELS_PER_UNIT", "Labels / Unit")
    sku = _first_value(row, "sku", "SKU", "item_number", "ITEM_NUMBER")

    label_required = "1" if in_packing_list else "0"
    if not case_qty:
        case_qty = _default_case_qty_for_product(packaging_level)
    if not labels_per_unit:
        labels_per_unit = _default_labels_per_unit_for_product(packaging_level)

    return {
        "id": _first_value(row, "id", "ROWID", "rowid"),
        "storefront": storefront,
        "in_packing_list": in_packing_list,
        "label_required": label_required,
        "gtin": gtin,
        "description": _first_value(row, "description", "DESCRIPTION", "Description"),
        "packaging_level": packaging_level,
        "dimensions_in": _first_value(row, "dimensions_in", "DIMENSIONS_IN", "L_X_W_X_H_IN", "L × W × H (in)"),
        "weight_lbs": _first_value(row, "weight_lbs", "WEIGHT_LBS", "Weight (lbs)"),
        "case_qty": case_qty,
        "labels_per_unit": labels_per_unit,
        "sku": sku,
        "unique_key": _product_master_unique_key(gtin, packaging_level, storefront, sku),
    }


def _product_master_unique_key(gtin: str, packaging_level: str, storefront: str = "", sku: str = "") -> str:
    store = _normalize_storefront(storefront)
    return "|".join([
        store.strip().lower(),
        normalize_packaging_level(packaging_level).strip().lower(),
        str(sku or "").strip().lower(),
    ])


def _product_storefront_level_sku_key(row: Dict[str, Any]) -> str:
    normalized = normalize_product_master_row(row)
    return _product_master_unique_key(
        normalized.get("gtin", ""),
        normalized.get("packaging_level", ""),
        normalized.get("storefront", ""),
        normalized.get("sku", ""),
    )


def _default_case_qty_for_product(packaging_level: str) -> str:
    return DEFAULT_CASE_QTY_BY_LEVEL.get(normalize_packaging_level(packaging_level), "")


def _default_labels_per_unit_for_product(packaging_level: str) -> str:
    return DEFAULT_LABELS_PER_UNIT_BY_LEVEL.get(normalize_packaging_level(packaging_level), "")


def parse_product_master_json(raw: Optional[str]) -> List[Dict[str, str]]:
    if not raw:
        return []
    try:
        data = json.loads(raw)
    except Exception:
        return []
    if not isinstance(data, list):
        return []
    return [normalize_product_master_row(r) for r in data if isinstance(r, dict)]


def _dedupe_product_master_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, str]]:
    deduped: Dict[str, Dict[str, str]] = {}
    fallback_index = 0
    for raw in rows:
        row = normalize_product_master_row(raw)
        has_data = any(row.get(k) for k in ("gtin", "description", "dimensions_in", "weight_lbs", "case_qty", "labels_per_unit", "sku"))
        if not has_data:
            continue
        key = row.get("unique_key") or ""
        if key == "|Other":
            fallback_index += 1
            key = f"row-{fallback_index}"
        deduped[key] = row
    return list(deduped.values())


def _kehe_product_master_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, str]]:
    return [row for row in _dedupe_product_master_rows(rows) if _is_kehe_storefront(row.get("storefront"))]


def _product_master_file_read(file_path: Optional[Path] = None) -> List[Dict[str, str]]:
    path = file_path or MPL_PRODUCT_MASTER_FILE
    try:
        if not path.exists():
            return []
        data = json.loads(path.read_text(encoding="utf-8"))
        rows = data.get("rows") if isinstance(data, dict) else data
        if not isinstance(rows, list):
            return []
        return _dedupe_product_master_rows([r for r in rows if isinstance(r, dict)])
    except Exception:
        return []


def _product_master_file_write(rows: List[Dict[str, Any]], file_path: Optional[Path] = None) -> List[Dict[str, str]]:
    path = file_path or MPL_PRODUCT_MASTER_FILE
    normalized = _dedupe_product_master_rows(rows)
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "rows": normalized,
        "updated_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
    }
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    return normalized


def _shared_product_master_file_read() -> List[Dict[str, str]]:
    return _product_master_file_read(MPL_PRODUCT_MASTER_FILE)


def _datastore_row_to_product(row: Dict[str, Any]) -> Dict[str, str]:
    return normalize_product_master_row(row)


def _product_to_datastore_row(row: Dict[str, Any], include_storefront: bool = False) -> Dict[str, Any]:
    normalized = normalize_product_master_row(row)
    out = {
        "LABEL_REQUIRED": "1" if normalized["in_packing_list"] else "0",
        "GTIN": normalized["gtin"],
        "DESCRIPTION": normalized["description"],
        "PACKAGING_LEVEL": normalized["packaging_level"],
        "DIMENSIONS_IN": normalized["dimensions_in"],
        "WEIGHT_LBS": normalized["weight_lbs"],
        "CASE_QTY": normalized["case_qty"],
        "LABELS_PER_UNIT": normalized["labels_per_unit"],
        "SKU": normalized["sku"],
        "UNIQUE_KEY": normalized["unique_key"],
        "IS_ACTIVE": True,
    }
    if include_storefront:
        out["STOREFRONT"] = normalized["storefront"]
    return out


def _init_catalyst_app(request: Request) -> Any:
    try:
        import zcatalyst_sdk  # type: ignore
    except Exception:
        return None
    try:
        return zcatalyst_sdk.initialize(scope="admin", req=request)
    except Exception:
        try:
            return zcatalyst_sdk.initialize(req=request)
        except Exception:
            return None


def _init_catalyst_user_app(request: Request) -> Any:
    try:
        import zcatalyst_sdk  # type: ignore
    except Exception:
        return None
    try:
        return zcatalyst_sdk.initialize(req=request)
    except Exception:
        return None


def _config_role_id_map() -> Dict[str, str]:
    raw = ACTIVE_LABELKIT_CONFIG.get("role_id_map", {})
    if not isinstance(raw, dict):
        return {}
    return {str(role_id).strip(): str(role).strip() for role_id, role in raw.items() if str(role_id).strip()}


def _role_from_role_id(role_id: str) -> str:
    mapped = _config_role_id_map().get(str(role_id or "").strip())
    return _role_from_name(mapped) if mapped else ""


def _role_from_name(role_name: str) -> str:
    normalized = str(role_name or "").strip().lower()
    if "admin" in normalized:
        return "Admin"
    if "editor" in normalized or "edit" in normalized:
        return "Editor"
    return "User"


def _role_from_catalyst(role_name: str = "", role_id: str = "") -> str:
    return _role_from_role_id(role_id) or _role_from_name(role_name)


def _request_user_from_headers(request: Request) -> Dict[str, Any]:
    headers = request.headers
    role_name = (
        headers.get("x-zc-user-role")
        or headers.get("x-zc-role-name")
        or headers.get("x-user-role")
        or headers.get("x-labelkit-role")
        or ""
    )
    role_id = (
        headers.get("x-zc-role-id")
        or headers.get("x-zc-user-role-id")
        or headers.get("x-user-role-id")
        or headers.get("x-labelkit-role-id")
        or ""
    )
    user = {
        "authenticated": False,
        "name": (
            headers.get("x-zc-user-name")
            or headers.get("x-user-name")
            or headers.get("x-forwarded-user")
            or headers.get("x-labelkit-user")
            or ""
        ),
        "email": (
            headers.get("x-zc-user-email")
            or headers.get("x-user-email")
            or headers.get("x-forwarded-email")
            or headers.get("x-labelkit-email")
            or ""
        ),
        "user_id": headers.get("x-zc-user-id") or headers.get("x-user-id") or "",
        "role": _role_from_catalyst(role_name, role_id),
        "role_name": role_name or "User",
        "role_id": role_id,
        "source": "headers",
    }
    # Catalyst may inject infrastructure IDs even before a real authenticated
    # user session exists, so do not treat a bare ID as signed in.
    user["authenticated"] = bool(user["email"] or user["name"])
    if not AUTH_REQUIRED and not user["authenticated"]:
        user.update({
            "authenticated": True,
            "name": "Local user",
            "email": "",
            "user_id": "",
            "role": "Admin",
            "role_name": "Local Admin",
            "source": "local",
        })
    return user


def _current_project_user(request: Request) -> Dict[str, Any]:
    cached = getattr(request.state, "labelkit_user", None)
    if isinstance(cached, dict):
        return cached

    user = _request_user_from_headers(request)
    catalyst_app = _init_catalyst_user_app(request)
    if catalyst_app is not None:
        try:
            details = catalyst_app.user_management().get_current_user()
            if isinstance(details, dict):
                role_details = details.get("role_details") or {}
                role_name = ""
                role_id = ""
                if isinstance(role_details, dict):
                    role_name = str(role_details.get("role_name") or "")
                    role_id = str(
                        role_details.get("role_id")
                        or role_details.get("roleId")
                        or role_details.get("id")
                        or ""
                    )
                name = " ".join([
                    str(details.get("first_name") or "").strip(),
                    str(details.get("last_name") or "").strip(),
                ]).strip()
                email = str(details.get("email_id") or details.get("email") or "")
                user_id = str(details.get("user_id") or details.get("zuid") or "")
                user = {
                    "authenticated": bool(email or name),
                    "name": name,
                    "email": email,
                    "user_id": user_id,
                    "role": _role_from_catalyst(role_name, role_id),
                    "role_name": role_name or "User",
                    "role_id": role_id,
                    "source": "catalyst",
                }
        except Exception:
            pass

    request.state.labelkit_user = user
    return user


def _permissions_for_role(role: str) -> Dict[str, bool]:
    role = _role_from_name(role)
    is_admin = role == "Admin"
    is_editor = role in {"Admin", "Editor"}
    return {
        "view": role in {"Admin", "Editor", "User"},
        "generate": role in {"Admin", "Editor", "User"},
        "table_crud": is_editor,
        "save_mpl": is_editor,
        "delete_mpl": is_admin,
        "audit_view": is_editor,
        "admin": is_admin,
    }


def _app_runtime_config(request: Optional[Request] = None) -> Dict[str, Any]:
    user = _current_project_user(request) if request is not None else {
        "authenticated": not AUTH_REQUIRED,
        "name": "Local user" if not AUTH_REQUIRED else "",
        "email": "",
        "user_id": "",
        "role": "Admin" if not AUTH_REQUIRED else "User",
        "role_name": "Local Admin" if not AUTH_REQUIRED else "User",
        "source": "local",
    }
    return {
        "app_env": APP_ENV,
        "config_profile": LABELKIT_CONFIG_PROFILE,
        "config_file": str(APP_CONFIG_FILE),
        "auth_required": AUTH_REQUIRED,
        "auth_mode": AUTH_MODE,
        "authenticated": bool(user.get("authenticated")),
        "allow_local_json_fallback": ALLOW_LOCAL_JSON_FALLBACK,
        "allow_browser_local_cache": ALLOW_BROWSER_LOCAL_CACHE,
        "user": user,
        "permissions": _permissions_for_role(str(user.get("role") or "User")),
        "store_modes": {
            "product_master": MPL_PRODUCT_MASTER_STORE,
            "directory": MPL_DIRECTORY_STORE,
            "mpl_drafts": MPL_DRAFTS_STORE,
            "audit_log": AUDIT_LOG_STORE,
        },
    }


@app.get("/api/auth/session")
async def auth_session(request: Request) -> JSONResponse:
    return JSONResponse(content=_app_runtime_config(request))


def _require_permission(request: Request, permission: str = "view") -> Dict[str, Any]:
    config = _app_runtime_config(request)
    user = config["user"]
    if AUTH_REQUIRED and not user.get("authenticated"):
        raise HTTPException(status_code=401, detail="Sign in with Catalyst Authentication to use LabelKit.")
    if not config["permissions"].get(permission, False):
        role_name = user.get("role_name") or user.get("role") or "User"
        raise HTTPException(status_code=403, detail=f"{role_name} does not have permission for this action.")
    return config


def _store_requires_datastore(store_mode: str) -> bool:
    return str(store_mode or "").strip().lower() == "datastore"


def _raise_datastore_unavailable(table_name: str, action: str, exc: Optional[Exception] = None) -> None:
    detail = (
        f"Cloud data unavailable for Catalyst Data Store table '{table_name}' while trying to {action}. "
        "Store mode is 'datastore', so local JSON fallback is disabled."
    )
    if exc is not None:
        detail += f" {exc.__class__.__name__}: {exc}"
    raise HTTPException(status_code=503, detail=detail)


def _datastore_table_named(request: Request, table_name: str, store_mode: str) -> Any:
    store_mode = str(store_mode or "auto").strip().lower()
    if store_mode == "file":
        return None

    catalyst_app = _init_catalyst_app(request)
    if catalyst_app is None:
        if _store_requires_datastore(store_mode):
            _raise_datastore_unavailable(table_name, "connect to it")
        return None

    try:
        return catalyst_app.datastore().table(table_name)
    except Exception as exc:
        if _store_requires_datastore(store_mode):
            _raise_datastore_unavailable(table_name, "open it", exc)
        return None


def _product_datastore_table(request: Request, table_name: str, store_mode: str) -> Any:
    return _datastore_table_named(request, table_name, store_mode)


def _datastore_get_raw_rows(table_service: Any) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    next_token: Optional[str] = None
    more_records = True

    while more_records:
        try:
            if next_token:
                page = table_service.get_paged_rows(next_token, max_rows=100)
            else:
                page = table_service.get_paged_rows(max_rows=100)
        except TypeError:
            page = table_service.get_paged_rows(next_token, 100)

        content = []
        if isinstance(page, dict):
            raw_content = page.get("content")
            if isinstance(raw_content, list):
                content = raw_content
            elif isinstance(page.get("data"), list):
                content = page.get("data", [])
        elif isinstance(page, list):
            content = page
        rows.extend([r for r in content if isinstance(r, dict)])
        more_records = bool(page.get("more_records")) if isinstance(page, dict) else False
        next_token = page.get("next_token") if isinstance(page, dict) else None
        if not next_token:
            more_records = False

    return rows


def _datastore_load_product_rows(request: Request, table_name: str, store_mode: str) -> Optional[List[Dict[str, str]]]:
    table_service = _product_datastore_table(request, table_name, store_mode)
    if table_service is None:
        return None
    try:
        raw_rows = _datastore_get_raw_rows(table_service)
        active_rows = [r for r in raw_rows if str(r.get("IS_ACTIVE", True)).lower() not in {"false", "0", "no"}]
        return _dedupe_product_master_rows([_datastore_row_to_product(r) for r in active_rows])
    except Exception as exc:
        if _store_requires_datastore(store_mode):
            _raise_datastore_unavailable(table_name, "read from it", exc)
        return None


def _datastore_load_product_master(request: Request) -> Optional[List[Dict[str, str]]]:
    return _datastore_load_product_rows(request, MPL_PRODUCT_MASTER_TABLE, MPL_PRODUCT_MASTER_STORE)


def _chunked(values: List[Any], size: int) -> List[List[Any]]:
    return [values[i:i + size] for i in range(0, len(values), size)]


def _now_iso() -> str:
    return time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())


def _request_actor(request: Request) -> Dict[str, str]:
    user = _current_project_user(request)
    name = str(user.get("name") or "")
    email = str(user.get("email") or "")
    user_id = str(user.get("user_id") or "")
    if not name and not email:
        name = "Local user"
    return {
        "name": name,
        "email": email,
        "user_id": user_id,
        "role": str(user.get("role") or ""),
        "role_name": str(user.get("role_name") or ""),
        "client": request.client.host if request.client else "",
    }


def _audit_datastore_table(request: Optional[Request]) -> Any:
    if request is None:
        if _store_requires_datastore(AUDIT_LOG_STORE):
            _raise_datastore_unavailable(AUDIT_LOG_TABLE, "connect to it")
        return None
    return _datastore_table_named(
        request,
        AUDIT_LOG_TABLE,
        AUDIT_LOG_STORE,
    )


def _datastore_row_to_audit(row: Dict[str, Any]) -> Dict[str, Any]:
    actor_raw = row.get("ACTOR_JSON") or row.get("actor") or {}
    actor = actor_raw if isinstance(actor_raw, dict) else {}
    if isinstance(actor_raw, str) and actor_raw.strip():
        try:
            parsed = json.loads(actor_raw)
            if isinstance(parsed, dict):
                actor = parsed
        except Exception:
            actor = {}
    return {
        "id": row.get("AUDIT_ID") or row.get("id") or row.get("ROWID") or "",
        "timestamp": row.get("EVENT_TIMESTAMP") or row.get("TIMESTAMP") or row.get("timestamp") or "",
        "actor": actor,
        "table": row.get("TABLE_NAME") or row.get("table") or "",
        "action": row.get("ACTION") or row.get("action") or "",
        "record_key": row.get("RECORD_KEY") or row.get("record_key") or "",
        "record_label": row.get("RECORD_LABEL") or row.get("record_label") or "",
        "field": row.get("FIELD_NAME") or row.get("field") or "",
        "old_value": row.get("OLD_VALUE") or row.get("old_value") or "",
        "new_value": row.get("NEW_VALUE") or row.get("new_value") or "",
        "source": row.get("SOURCE") or row.get("source") or "",
        "batch_id": row.get("BATCH_ID") or row.get("batch_id") or "",
        "filename": row.get("FILENAME") or row.get("filename") or "",
    }


def _audit_to_datastore_row(entry: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "AUDIT_ID": str(entry.get("id") or uuid.uuid4().hex),
        "EVENT_TIMESTAMP": str(entry.get("timestamp") or _now_iso()),
        "ACTOR_JSON": json.dumps(entry.get("actor") or {}, sort_keys=True),
        "TABLE_NAME": str(entry.get("table") or ""),
        "ACTION": str(entry.get("action") or ""),
        "RECORD_KEY": str(entry.get("record_key") or ""),
        "RECORD_LABEL": str(entry.get("record_label") or ""),
        "FIELD_NAME": str(entry.get("field") or ""),
        "OLD_VALUE": str(entry.get("old_value") or ""),
        "NEW_VALUE": str(entry.get("new_value") or ""),
        "SOURCE": str(entry.get("source") or ""),
        "BATCH_ID": str(entry.get("batch_id") or ""),
        "FILENAME": str(entry.get("filename") or ""),
        "IS_ACTIVE": True,
    }


def _datastore_load_audit_log(request: Optional[Request], limit: Optional[int] = None, table: str = "") -> Optional[List[Dict[str, Any]]]:
    table_service = _audit_datastore_table(request)
    if table_service is None:
        return None
    try:
        raw_rows = _datastore_get_raw_rows(table_service)
        active_rows = [
            row for row in raw_rows
            if str(row.get("IS_ACTIVE", True)).lower() not in {"false", "0", "no"}
        ]
        rows = [_datastore_row_to_audit(row) for row in active_rows]
        if table:
            rows = [row for row in rows if str(row.get("table", "")) == table]
        rows.sort(key=lambda row: str(row.get("timestamp", "")), reverse=True)
        return rows[:limit] if limit else rows
    except Exception as exc:
        if _store_requires_datastore(AUDIT_LOG_STORE):
            _raise_datastore_unavailable(AUDIT_LOG_TABLE, "read from it", exc)
        return None


def _datastore_append_audit_log(request: Optional[Request], entries: List[Dict[str, Any]]) -> bool:
    table_service = _audit_datastore_table(request)
    if table_service is None:
        return False
    try:
        insert_rows = [_audit_to_datastore_row(entry) for entry in entries]
        for batch in _chunked(insert_rows, 100):
            if batch:
                table_service.insert_rows(batch)
        return True
    except Exception as exc:
        if _store_requires_datastore(AUDIT_LOG_STORE):
            _raise_datastore_unavailable(AUDIT_LOG_TABLE, "write to it", exc)
        return False


def _audit_log_read(limit: Optional[int] = None, table: str = "", request: Optional[Request] = None) -> List[Dict[str, Any]]:
    datastore_rows = _datastore_load_audit_log(request, limit=limit, table=table)
    if datastore_rows is not None:
        return datastore_rows
    try:
        if not AUDIT_LOG_FILE.exists():
            return []
        data = json.loads(AUDIT_LOG_FILE.read_text(encoding="utf-8"))
        entries = data.get("entries") if isinstance(data, dict) else data
        if not isinstance(entries, list):
            return []
        rows = [entry for entry in entries if isinstance(entry, dict)]
        if table:
            rows = [row for row in rows if str(row.get("table", "")) == table]
        rows.sort(key=lambda row: str(row.get("timestamp", "")), reverse=True)
        return rows[:limit] if limit else rows
    except Exception:
        return []


def _audit_log_append(entries: List[Dict[str, Any]], request: Optional[Request] = None) -> None:
    if not entries:
        return
    if _datastore_append_audit_log(request, entries):
        return
    AUDIT_LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
    existing = _audit_log_read()
    combined = entries + existing
    payload = {
        "entries": combined[:1000],
        "updated_at": _now_iso(),
    }
    AUDIT_LOG_FILE.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def _audit_value(value: Any) -> str:
    if isinstance(value, (list, dict)):
        return json.dumps(value, sort_keys=True)
    return str(value if value is not None else "")


def _product_row_key(row: Dict[str, Any]) -> str:
    normalized = normalize_product_master_row(row)
    return _product_storefront_level_sku_key(normalized)


def _dc_row_key(row: Dict[str, Any]) -> str:
    normalized = normalize_dc_directory_row(row)
    return _dc_directory_base_key(normalized.get("dc", ""), normalized.get("storefront", ""))


def _row_label(table: str, row: Dict[str, Any]) -> str:
    if table in {"kehe_product_master", "mpl_product_master"}:
        normalized = normalize_product_master_row(row)
        return normalized.get("description") or normalized.get("gtin") or normalized.get("sku") or "Product row"
    if table in {"kehe_dc_directory", "mpl_directory"}:
        normalized = normalize_dc_directory_row(row)
        return normalized.get("name") or normalized.get("dc") or "DC row"
    if table == "kehe_mpl_drafts":
        return str(row.get("name") or row.get("id") or "MPL draft")
    return "Record"


def _normalized_for_audit(table: str, row: Dict[str, Any]) -> Dict[str, Any]:
    if table in {"kehe_product_master", "mpl_product_master"}:
        normalized = normalize_product_master_row(row)
        normalized.pop("unique_key", None)
        return normalized
    if table in {"kehe_dc_directory", "mpl_directory"}:
        normalized = normalize_dc_directory_row(row)
        normalized.pop("unique_key", None)
        normalized.pop("id", None)
        return normalized
    return dict(row)


def _audit_row_changes(
    *,
    request: Request,
    table: str,
    old_rows: List[Dict[str, Any]],
    new_rows: List[Dict[str, Any]],
    key_fn: Any,
    source: str,
    batch_id: str = "",
    filename: str = "",
) -> None:
    actor = _request_actor(request)
    timestamp = _now_iso()
    old_by_key = {key_fn(row): _normalized_for_audit(table, row) for row in old_rows}
    new_by_key = {key_fn(row): _normalized_for_audit(table, row) for row in new_rows}
    labels = {key_fn(row): _row_label(table, row) for row in [*old_rows, *new_rows]}
    entries: List[Dict[str, Any]] = []

    for key in sorted(set(old_by_key) | set(new_by_key)):
        old = old_by_key.get(key)
        new = new_by_key.get(key)
        if old is None and new is not None:
            entries.append({
                "id": uuid.uuid4().hex,
                "timestamp": timestamp,
                "actor": actor,
                "table": table,
                "action": "create",
                "record_key": key,
                "record_label": labels.get(key, ""),
                "field": "__row__",
                "old_value": "",
                "new_value": _audit_value(new),
                "source": source,
                "batch_id": batch_id,
                "filename": filename,
            })
            continue
        if old is not None and new is None:
            entries.append({
                "id": uuid.uuid4().hex,
                "timestamp": timestamp,
                "actor": actor,
                "table": table,
                "action": "delete",
                "record_key": key,
                "record_label": labels.get(key, ""),
                "field": "__row__",
                "old_value": _audit_value(old),
                "new_value": "",
                "source": source,
                "batch_id": batch_id,
                "filename": filename,
            })
            continue
        if old is None or new is None:
            continue
        for field in sorted(set(old) | set(new)):
            old_value = _audit_value(old.get(field))
            new_value = _audit_value(new.get(field))
            if old_value == new_value:
                continue
            entries.append({
                "id": uuid.uuid4().hex,
                "timestamp": timestamp,
                "actor": actor,
                "table": table,
                "action": "update",
                "record_key": key,
                "record_label": labels.get(key, ""),
                "field": field,
                "old_value": old_value,
                "new_value": new_value,
                "source": source,
                "batch_id": batch_id,
                "filename": filename,
            })

    _audit_log_append(entries, request=request)


def _datastore_save_product_rows(
    request: Request,
    rows: List[Dict[str, Any]],
    table_name: str,
    store_mode: str,
    *,
    include_storefront: bool = False,
) -> Optional[List[Dict[str, str]]]:
    table_service = _product_datastore_table(request, table_name, store_mode)
    if table_service is None:
        return None
    normalized = _dedupe_product_master_rows(rows)
    try:
        existing_rows = _datastore_get_raw_rows(table_service)
        row_ids = [r.get("ROWID") for r in existing_rows if r.get("ROWID")]
        for batch in _chunked(row_ids, 200):
            table_service.delete_rows(batch)
        insert_rows = [_product_to_datastore_row(r, include_storefront=include_storefront) for r in normalized]
        for batch in _chunked(insert_rows, 100):
            if batch:
                table_service.insert_rows(batch)
        return normalized
    except Exception as exc:
        if _store_requires_datastore(store_mode):
            _raise_datastore_unavailable(table_name, "write to it", exc)
        return None


def _datastore_save_product_master(request: Request, rows: List[Dict[str, Any]]) -> Optional[List[Dict[str, str]]]:
    return _datastore_save_product_rows(
        request,
        _dedupe_product_master_rows(rows),
        MPL_PRODUCT_MASTER_TABLE,
        MPL_PRODUCT_MASTER_STORE,
        include_storefront=True,
    )


@app.get("/api/kehe/product-master")
async def get_kehe_product_master(request: Request) -> JSONResponse:
    _require_permission(request, "view")
    rows = _datastore_load_product_master(request)
    source = "datastore"
    if rows is None:
        rows = _shared_product_master_file_read()
        source = "file"
    rows = _kehe_product_master_rows(rows)
    return JSONResponse(content={"rows": rows, "source": source})


@app.put("/api/kehe/product-master")
async def save_kehe_product_master(request: Request, payload: Dict[str, Any]) -> JSONResponse:
    raise HTTPException(
        status_code=405,
        detail="KeHE Product Master is read-only. Edit shared rows from Packing List & Ti-Hi Product Master.",
    )


@app.get("/api/mpl/product-master")
async def get_mpl_product_master(request: Request) -> JSONResponse:
    _require_permission(request, "view")
    rows = _datastore_load_product_master(request)
    source = "datastore"
    if rows is None:
        rows = _shared_product_master_file_read()
        source = "file"
        if rows and not MPL_PRODUCT_MASTER_FILE.exists():
            rows = _product_master_file_write(rows, MPL_PRODUCT_MASTER_FILE)
    return JSONResponse(content={"rows": rows, "source": source})


@app.put("/api/mpl/product-master")
async def save_mpl_product_master(request: Request, payload: Dict[str, Any]) -> JSONResponse:
    _require_permission(request, "table_crud")
    rows = payload.get("rows") if isinstance(payload, dict) else []
    if not isinstance(rows, list):
        rows = []
    rows = _dedupe_product_master_rows(rows)
    source = str(payload.get("source") or "manual_edit") if isinstance(payload, dict) else "manual_edit"
    batch_id = str(payload.get("batch_id") or "") if isinstance(payload, dict) else ""
    filename = str(payload.get("filename") or "") if isinstance(payload, dict) else ""

    old_rows = _datastore_load_product_master(request)
    if old_rows is None:
        old_rows = _shared_product_master_file_read()

    saved_rows = _datastore_save_product_master(request, rows)
    storage_source = "datastore"
    if saved_rows is None:
        saved_rows = _product_master_file_write(rows, MPL_PRODUCT_MASTER_FILE)
        storage_source = "file"

    _audit_row_changes(
        request=request,
        table="mpl_product_master",
        old_rows=old_rows,
        new_rows=saved_rows,
        key_fn=_product_row_key,
        source=source,
        batch_id=batch_id,
        filename=filename,
    )

    return JSONResponse(content={"rows": saved_rows, "saved": True, "source": storage_source})


# ---------------------------------------------------------------------------
# BACKEND SECTION 4D: KeHE DC Directory persistence.
# Frontend uses these APIs for the editable DC Directory modal.
# ---------------------------------------------------------------------------
def _parse_match_values(value: Any) -> List[str]:
    if isinstance(value, list):
        return [str(v).strip() for v in value if str(v).strip()]

    raw = str(value or "").strip()
    if not raw:
        return []

    try:
        parsed = json.loads(raw)
        if isinstance(parsed, list):
            return [str(v).strip() for v in parsed if str(v).strip()]
    except Exception:
        pass

    return [v.strip() for v in re.split(r"[\n,]+", raw) if v.strip()]


def normalize_dc_directory_row(row: Dict[str, Any]) -> Dict[str, Any]:
    storefront = _normalize_storefront(_first_value(row, "storefront", "STOREFRONT", "Storefront"))
    dc = _first_value(row, "dc", "DC")
    name = _first_value(row, "name", "NAME")
    ship_from = _first_value(row, "ship_from", "SHIP_FROM", "ship_from_address", "SHIP_FROM_ADDRESS", "Ship From")
    delivery_address = _first_value(row, "delivery_address", "DELIVERY_ADDRESS")
    billing_address = _first_value(row, "billing_address", "BILLING_ADDRESS")
    match_values = _parse_match_values(row.get("match_values", row.get("MATCH_VALUES", [])))
    return {
        "id": _first_value(row, "id", "ROWID", "rowid"),
        "storefront": storefront,
        "dc": dc,
        "name": name,
        "ship_from": ship_from or DEFAULT_KEHE_SHIP_FROM,
        "delivery_address": delivery_address,
        "billing_address": billing_address,
        "match_values": match_values,
        "unique_key": _dc_directory_unique_key(
            dc,
            storefront,
            name,
            delivery_address,
            billing_address,
            match_values,
        ),
    }


def _dc_directory_base_key(dc: str, storefront: str = "") -> str:
    store = _normalize_storefront(storefront)
    return f"{store.strip().lower()}|{str(dc or '').strip().lower()}"


def _dc_directory_unique_key(
    dc: str,
    storefront: str = "",
    name: str = "",
    delivery_address: str = "",
    billing_address: str = "",
    match_values: Optional[List[str]] = None,
) -> str:
    return _dc_directory_base_key(dc, storefront)


def _dedupe_dc_directory_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    deduped: Dict[str, Dict[str, Any]] = {}
    fallback_index = 0

    for raw in rows:
        row = normalize_dc_directory_row(raw)
        if not any([row.get("dc"), row.get("name"), row.get("ship_from"), row.get("delivery_address"), row.get("billing_address"), row.get("match_values")]):
            continue

        key = str(row.get("unique_key") or row.get("dc") or "").strip()
        if not key:
            fallback_index += 1
            key = f"row-{fallback_index}"

        deduped[key] = row

    return list(deduped.values())


def _kehe_dc_directory_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return [row for row in _dedupe_dc_directory_rows(rows) if _is_kehe_storefront(row.get("storefront"))]


def _dc_directory_file_read(file_path: Optional[Path] = None) -> List[Dict[str, Any]]:
    path = file_path or MPL_DIRECTORY_FILE
    try:
        if not path.exists():
            return []

        data = json.loads(path.read_text(encoding="utf-8"))

        if isinstance(data, dict):
            if isinstance(data.get("rows"), list):
                return _dedupe_dc_directory_rows([r for r in data.get("rows", []) if isinstance(r, dict)])
            rows = []
            for dc, row in data.items():
                if isinstance(row, dict):
                    merged = {"dc": dc, **row}
                    rows.append(merged)
            return _dedupe_dc_directory_rows(rows)

        if isinstance(data, list):
            return _dedupe_dc_directory_rows([r for r in data if isinstance(r, dict)])

        return []
    except Exception:
        return []


def _dc_directory_file_write(
    rows: List[Dict[str, Any]],
    file_path: Optional[Path] = None,
) -> List[Dict[str, Any]]:
    path = file_path or MPL_DIRECTORY_FILE
    normalized = _dedupe_dc_directory_rows(rows)

    path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "rows": normalized,
        "updated_at": _now_iso(),
    }
    path.write_text(
        json.dumps(payload, indent=2),
        encoding="utf-8",
    )

    try:
        load_kehe_dc_directory.cache_clear()
    except Exception:
        pass

    return normalized


def _shared_dc_directory_file_read() -> List[Dict[str, Any]]:
    return _dc_directory_file_read(MPL_DIRECTORY_FILE)


def _datastore_row_to_dc(row: Dict[str, Any]) -> Dict[str, Any]:
    return normalize_dc_directory_row(row)


def _dc_to_datastore_row(row: Dict[str, Any], include_storefront: bool = False) -> Dict[str, Any]:
    normalized = normalize_dc_directory_row(row)

    out = {
        "DC": normalized["dc"],
        "NAME": normalized["name"],
        "SHIP_FROM": normalized["ship_from"],
        "DELIVERY_ADDRESS": normalized["delivery_address"],
        "BILLING_ADDRESS": normalized["billing_address"],
        "MATCH_VALUES": json.dumps(normalized["match_values"]),
        "UNIQUE_KEY": normalized["unique_key"],
        "IS_ACTIVE": True,
    }
    if include_storefront:
        out["STOREFRONT"] = normalized["storefront"]
    return out


def _dc_datastore_table_named(request: Request, table_name: str, store_mode: str) -> Any:
    return _datastore_table_named(request, table_name, store_mode)


def _datastore_load_dc_rows(request: Request, table_name: str, store_mode: str) -> Optional[List[Dict[str, Any]]]:
    table_service = _dc_datastore_table_named(request, table_name, store_mode)
    if table_service is None:
        return None

    try:
        raw_rows = _datastore_get_raw_rows(table_service)
        active_rows = [
            r for r in raw_rows
            if str(r.get("IS_ACTIVE", True)).lower() not in {"false", "0", "no"}
        ]
        return _dedupe_dc_directory_rows([_datastore_row_to_dc(r) for r in active_rows])
    except Exception as exc:
        if _store_requires_datastore(store_mode):
            _raise_datastore_unavailable(table_name, "read from it", exc)
        return None


def _datastore_load_dc_directory(request: Request) -> Optional[List[Dict[str, Any]]]:
    return _datastore_load_dc_rows(request, MPL_DIRECTORY_TABLE, MPL_DIRECTORY_STORE)


def _datastore_save_dc_rows(
    request: Request,
    rows: List[Dict[str, Any]],
    table_name: str,
    store_mode: str,
    *,
    include_storefront: bool = False,
) -> Optional[List[Dict[str, Any]]]:
    table_service = _dc_datastore_table_named(request, table_name, store_mode)
    if table_service is None:
        return None

    normalized = _dedupe_dc_directory_rows(rows)

    try:
        existing_rows = _datastore_get_raw_rows(table_service)
        row_ids = [r.get("ROWID") for r in existing_rows if r.get("ROWID")]

        for batch in _chunked(row_ids, 200):
            table_service.delete_rows(batch)

        insert_rows = [_dc_to_datastore_row(r, include_storefront=include_storefront) for r in normalized]

        for batch in _chunked(insert_rows, 100):
            if batch:
                table_service.insert_rows(batch)

        return normalized
    except Exception as exc:
        if _store_requires_datastore(store_mode):
            _raise_datastore_unavailable(table_name, "write to it", exc)
        return None


def _datastore_save_dc_directory(request: Request, rows: List[Dict[str, Any]]) -> Optional[List[Dict[str, Any]]]:
    return _datastore_save_dc_rows(
        request,
        _dedupe_dc_directory_rows(rows),
        MPL_DIRECTORY_TABLE,
        MPL_DIRECTORY_STORE,
        include_storefront=True,
    )


def _sync_kehe_dc_directory_for_pipeline(request: Request) -> List[Dict[str, Any]]:
    rows = _datastore_load_dc_directory(request)

    if rows is None:
        rows = _shared_dc_directory_file_read()
    elif not _store_requires_datastore(MPL_DIRECTORY_STORE):
        _dc_directory_file_write(rows, MPL_DIRECTORY_FILE)

    rows = _kehe_dc_directory_rows(rows)
    try:
        load_kehe_dc_directory.cache_clear()
    except Exception:
        pass

    return rows


@app.get("/api/kehe/dc-directory")
async def get_kehe_dc_directory(request: Request) -> JSONResponse:
    _require_permission(request, "view")
    rows = _datastore_load_dc_directory(request)
    source = "datastore"

    if rows is None:
        rows = _shared_dc_directory_file_read()
        source = "file"
    rows = _kehe_dc_directory_rows(rows)

    return JSONResponse(content={"rows": rows, "source": source})


@app.put("/api/kehe/dc-directory")
async def save_kehe_dc_directory(request: Request, payload: Dict[str, Any]) -> JSONResponse:
    raise HTTPException(
        status_code=405,
        detail="KeHE DC Directory is read-only. Edit shared rows from Packing List & Ti-Hi Directory.",
    )


@app.get("/api/mpl/directory")
async def get_mpl_directory(request: Request) -> JSONResponse:
    _require_permission(request, "view")
    rows = _datastore_load_dc_directory(request)
    source = "datastore"

    if rows is None:
        rows = _shared_dc_directory_file_read()
        source = "file"
        if rows and not MPL_DIRECTORY_FILE.exists():
            rows = _dc_directory_file_write(
                rows,
                MPL_DIRECTORY_FILE,
            )

    return JSONResponse(content={"rows": rows, "source": source})


@app.put("/api/mpl/directory")
async def save_mpl_directory(request: Request, payload: Dict[str, Any]) -> JSONResponse:
    _require_permission(request, "table_crud")
    rows = payload.get("rows") if isinstance(payload, dict) else []

    if not isinstance(rows, list):
        rows = []
    source = str(payload.get("source") or "manual_edit") if isinstance(payload, dict) else "manual_edit"
    batch_id = str(payload.get("batch_id") or "") if isinstance(payload, dict) else ""
    filename = str(payload.get("filename") or "") if isinstance(payload, dict) else ""

    rows = _dedupe_dc_directory_rows(rows)
    old_rows = _datastore_load_dc_directory(request)
    if old_rows is None:
        old_rows = _shared_dc_directory_file_read()

    saved_rows = _datastore_save_dc_directory(request, rows)
    storage_source = "datastore"

    if saved_rows is None:
        saved_rows = _dc_directory_file_write(
            rows,
            MPL_DIRECTORY_FILE,
        )
        storage_source = "file"

    _audit_row_changes(
        request=request,
        table="mpl_directory",
        old_rows=old_rows,
        new_rows=saved_rows,
        key_fn=_dc_row_key,
        source=source,
        batch_id=batch_id,
        filename=filename,
    )

    return JSONResponse(content={"rows": saved_rows, "saved": True, "source": storage_source})


@app.get("/api/kehe/audit-log")
async def get_kehe_audit_log(request: Request, limit: int = 200, table: str = "") -> JSONResponse:
    _require_permission(request, "audit_view")
    safe_limit = min(max(int(limit or 200), 1), 1000)
    return JSONResponse(content={"entries": _audit_log_read(limit=safe_limit, table=table, request=request)})


class _ZohoAnalyticsRequestError(Exception):
    def __init__(self, status_code: int, message: str):
        super().__init__(message)
        self.status_code = status_code
        self.message = message


def _analytics_connection_details(request: Request) -> tuple[Dict[str, str], Dict[str, str]]:
    catalyst_app = _init_catalyst_app(request)
    if catalyst_app is None:
        raise HTTPException(
            status_code=503,
            detail=(
                "Zoho Analytics order lookup is available from the deployed Catalyst AppSail app. "
                "The Catalyst SDK or runtime connection is unavailable here."
            ),
        )

    connections_factory = getattr(catalyst_app, "connections", None)
    if not callable(connections_factory):
        raise HTTPException(
            status_code=503,
            detail="The installed Catalyst SDK does not support Connections. Install zcatalyst-sdk 1.1 or newer.",
        )

    try:
        response = connections_factory().get_connection_credentials(ANALYTICS_CONNECTION_LINK_NAME)
    except Exception as exc:
        raise HTTPException(
            status_code=503,
            detail=(
                f"Catalyst Connection '{ANALYTICS_CONNECTION_LINK_NAME}' could not provide Zoho Analytics credentials. "
                "Confirm that the connection is active, shared with this environment, and has ZohoAnalytics.data.read. "
                f"{exc.__class__.__name__}: {exc}"
            ),
        ) from exc

    payload = response if isinstance(response, dict) else {}
    details = payload.get("connections") if isinstance(payload.get("connections"), dict) else payload
    raw_headers = details.get("headers") if isinstance(details, dict) else {}
    raw_parameters = details.get("parameters") if isinstance(details, dict) else {}
    headers = {
        str(key): str(value)
        for key, value in (raw_headers.items() if isinstance(raw_headers, dict) else [])
        if str(key).strip() and value is not None
    }
    parameters = {
        str(key): str(value)
        for key, value in (raw_parameters.items() if isinstance(raw_parameters, dict) else [])
        if str(key).strip() and value is not None
    }
    if not headers and not parameters:
        raise HTTPException(
            status_code=503,
            detail=f"Catalyst Connection '{ANALYTICS_CONNECTION_LINK_NAME}' returned no authentication details.",
        )
    return headers, parameters


def _analytics_header_value(headers: Dict[str, str], name: str) -> str:
    wanted = name.strip().lower()
    for key, value in headers.items():
        if str(key).strip().lower() == wanted:
            return str(value or "").strip()
    return ""


def _analytics_url(path: str, parameters: Optional[Dict[str, str]] = None) -> str:
    url = f"{ANALYTICS_API_BASE}/{str(path or '').lstrip('/')}"
    if parameters:
        url += "?" + urllib.parse.urlencode(parameters)
    return url


def _analytics_error_detail(raw_body: bytes, fallback: str) -> str:
    body = raw_body.decode("utf-8", errors="replace").strip()
    if body:
        try:
            payload = json.loads(body)
            data = payload.get("data") if isinstance(payload, dict) else {}
            if isinstance(data, dict):
                message = data.get("errorMessage") or data.get("message")
                if message:
                    return str(message)[:600]
            if isinstance(payload, dict) and (payload.get("summary") or payload.get("message")):
                return str(payload.get("summary") or payload.get("message"))[:600]
        except Exception:
            pass
        return body[:600]
    return fallback


def _analytics_http_get(url: str, headers: Dict[str, str]) -> bytes:
    request_headers = {"Accept": "application/json, text/csv;q=0.9", **headers}
    api_request = urllib.request.Request(url, headers=request_headers, method="GET")
    try:
        with urllib.request.urlopen(api_request, timeout=ANALYTICS_HTTP_TIMEOUT_SECONDS) as response:
            return response.read()
    except urllib.error.HTTPError as exc:
        raw_body = exc.read() if exc.fp is not None else b""
        raise _ZohoAnalyticsRequestError(
            int(exc.code or 502),
            _analytics_error_detail(raw_body, str(exc.reason or "Zoho Analytics request failed.")),
        ) from exc
    except urllib.error.URLError as exc:
        raise _ZohoAnalyticsRequestError(502, f"Zoho Analytics is unavailable: {exc.reason}") from exc
    except TimeoutError as exc:
        raise _ZohoAnalyticsRequestError(504, "Zoho Analytics did not respond before the request timed out.") from exc


def _analytics_discover_org_ids(headers: Dict[str, str], parameters: Dict[str, str]) -> List[str]:
    try:
        raw = _analytics_http_get(
            _analytics_url("/restapi/v2/orgs", parameters),
            headers,
        )
        payload = json.loads(raw.decode("utf-8-sig", errors="replace"))
        data = payload.get("data") if isinstance(payload, dict) else {}
        orgs = data.get("orgs") if isinstance(data, dict) else []
        if not isinstance(orgs, list):
            return []
        default_ids = [
            str(org.get("orgId") or "").strip()
            for org in orgs
            if isinstance(org, dict) and org.get("isDefault") and str(org.get("orgId") or "").strip()
        ]
        other_ids = [
            str(org.get("orgId") or "").strip()
            for org in orgs
            if isinstance(org, dict) and not org.get("isDefault") and str(org.get("orgId") or "").strip()
        ]
        return default_ids + other_ids
    except _ZohoAnalyticsRequestError as exc:
        raise HTTPException(
            status_code=503,
            detail=(
                "The Zoho Analytics organization ID is not configured and could not be discovered. "
                "Set analytics_org_id in labelkit_config.json, or add ZohoAnalytics.metadata.read to the "
                f"'{ANALYTICS_CONNECTION_LINK_NAME}' connection. {exc.message}"
            ),
        ) from exc
    except (TypeError, ValueError, json.JSONDecodeError) as exc:
        raise HTTPException(
            status_code=502,
            detail="Zoho Analytics returned an invalid organization response.",
        ) from exc


def _analytics_export_order_rows(request: Request, sales_order_number: str) -> List[Dict[str, str]]:
    global ANALYTICS_DISCOVERED_ORG_ID

    if ANALYTICS_ORDER_SOURCE in ANALYTICS_LOCAL_SOURCE_VALUES:
        if ANALYTICS_LOCAL_FILE is None:
            raise HTTPException(
                status_code=503,
                detail="Local Analytics testing requires 'analytics_local_file' in the local LabelKit profile.",
            )
        if not ANALYTICS_LOCAL_FILE.is_file():
            raise HTTPException(
                status_code=503,
                detail=(
                    f"Local order workbook was not found at '{ANALYTICS_LOCAL_FILE}'. "
                    "Add the configured .csv or .xlsx file before searching for an order."
                ),
            )
        try:
            local_rows = _read_spreadsheet_bytes(
                ANALYTICS_LOCAL_FILE.name,
                ANALYTICS_LOCAL_FILE.read_bytes(),
            )
        except OSError as exc:
            raise HTTPException(
                status_code=500,
                detail=f"Local order workbook could not be read: {exc}",
            ) from exc
        wanted_order = _canonical_order_number(sales_order_number)
        matching_rows = [
            {str(key): value for key, value in row.items()}
            for row in local_rows
            if _canonical_order_number(_analytics_row_value(row, ANALYTICS_ORDER_COLUMN)) == wanted_order
        ]
        return matching_rows

    headers, connection_parameters = _analytics_connection_details(request)
    header_org_id = _analytics_header_value(headers, "ZANALYTICS-ORGID")
    configured_org_id = ANALYTICS_ORG_ID or header_org_id or ANALYTICS_DISCOVERED_ORG_ID
    org_ids = [configured_org_id] if configured_org_id else _analytics_discover_org_ids(headers, connection_parameters)
    org_ids = list(dict.fromkeys(org_id for org_id in org_ids if org_id))
    if not org_ids:
        raise HTTPException(
            status_code=503,
            detail="No accessible Zoho Analytics organization was found for the 'orderdata' connection.",
        )

    safe_order = sales_order_number.replace("'", "''")
    safe_order_column = ANALYTICS_ORDER_COLUMN.replace('"', '""')
    export_config = {
        "responseFormat": "csv",
        "criteria": f'"{safe_order_column}"=\'{safe_order}\'',
        "selectedColumns": list(dict.fromkeys([
            ANALYTICS_ORDER_COLUMN,
            ANALYTICS_SKU_COLUMN,
            ANALYTICS_QUANTITY_COLUMN,
            ANALYTICS_ORDER_INSTANCE_ID_COLUMN,
            ANALYTICS_ORDER_INSTANCE_DATE_COLUMN,
            ANALYTICS_ORDER_INSTANCE_STOREFRONT_COLUMN,
            *ANALYTICS_ORDER_DETAIL_COLUMNS.values(),
        ])),
        "includeHeader": True,
    }
    parameters = dict(connection_parameters)
    parameters["CONFIG"] = json.dumps(export_config, separators=(",", ":"))
    path = f"/restapi/v2/workspaces/{ANALYTICS_WORKSPACE_ID}/views/{ANALYTICS_VIEW_ID}/data"
    last_error: Optional[_ZohoAnalyticsRequestError] = None

    for org_id in org_ids:
        request_headers = {
            key: value
            for key, value in headers.items()
            if str(key).strip().lower() != "zanalytics-orgid"
        }
        request_headers["ZANALYTICS-ORGID"] = org_id
        try:
            raw = _analytics_http_get(_analytics_url(path, parameters), request_headers)
            ANALYTICS_DISCOVERED_ORG_ID = org_id
            text = raw.decode("utf-8-sig", errors="replace")
            wanted_order = _canonical_order_number(sales_order_number)
            matching_rows = [
                dict(row)
                for row in csv.DictReader(io.StringIO(text))
                if _canonical_order_number(_analytics_row_value(row, ANALYTICS_ORDER_COLUMN)) == wanted_order
            ]
            return matching_rows
        except _ZohoAnalyticsRequestError as exc:
            last_error = exc

    error_detail = last_error.message if last_error else "The Analytics view could not be read."
    raise HTTPException(
        status_code=502,
        detail=(
            f"Zoho Analytics could not read '{ANALYTICS_VIEW_NAME}' with connection "
            f"'{ANALYTICS_CONNECTION_LINK_NAME}'. {error_detail}"
        ),
    )


def _analytics_row_value(row: Dict[str, Any], column_name: str) -> str:
    wanted = str(column_name or "").strip().lower()
    for key, value in row.items():
        if str(key or "").strip().lower() == wanted:
            return str(value or "").strip()
    return ""


def _canonical_order_sku(value: Any) -> str:
    raw = str(value or "").strip().lower()
    if re.fullmatch(r"\d+(?:\.0+)?", raw):
        raw = raw.split(".", 1)[0].lstrip("0") or "0"
    return re.sub(r"[\s_-]+", "", raw)


def _canonical_order_number(value: Any) -> str:
    raw = str(value or "").strip()
    if re.fullmatch(r"\d+\.0+", raw):
        raw = raw.split(".", 1)[0]
    return raw.casefold()


def _analytics_order_instance_groups(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    groups: Dict[str, Dict[str, Any]] = {}
    group_order: List[str] = []
    for row in rows:
        raw_ecomdash_id = _analytics_row_value(row, ANALYTICS_ORDER_INSTANCE_ID_COLUMN)
        canonical_ecomdash_id = _canonical_order_number(raw_ecomdash_id)
        invoice_date = _analytics_row_value(row, ANALYTICS_ORDER_INSTANCE_DATE_COLUMN)
        storefront = _analytics_row_value(row, ANALYTICS_ORDER_INSTANCE_STOREFRONT_COLUMN)
        billing_customer_name = _analytics_row_value(row, "Billing Customer Name")
        if canonical_ecomdash_id:
            group_key = f"ecomdash:{canonical_ecomdash_id}"
        else:
            group_key = "missing:" + "|".join([
                invoice_date.casefold(),
                storefront.casefold(),
                billing_customer_name.casefold(),
            ])
        if group_key not in groups:
            groups[group_key] = {
                "key": group_key,
                "ecomdash_id": raw_ecomdash_id,
                "storefront": storefront,
                "billing_customer_name": billing_customer_name,
                "invoice_date": invoice_date,
                "rows": [],
            }
            group_order.append(group_key)
        groups[group_key]["rows"].append(row)

    instances: List[Dict[str, Any]] = []
    for group_key in group_order:
        group = groups[group_key]
        sku_keys = {
            _canonical_order_sku(_analytics_row_value(row, ANALYTICS_SKU_COLUMN))
            for row in group["rows"]
            if _canonical_order_sku(_analytics_row_value(row, ANALYTICS_SKU_COLUMN))
        }
        group["line_count"] = len(group["rows"])
        group["sku_count"] = len(sku_keys)
        instances.append(group)
    return instances


def _analytics_order_instance_summary(instance: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "ecomdash_id": instance.get("ecomdash_id", ""),
        "storefront": instance.get("storefront", ""),
        "billing_customer_name": instance.get("billing_customer_name", ""),
        "invoice_date": instance.get("invoice_date", ""),
        "line_count": instance.get("line_count", 0),
        "sku_count": instance.get("sku_count", 0),
    }


def _analytics_quantity(value: Any) -> Optional[float | int]:
    raw = str(value or "").strip().replace(",", "")
    if not raw:
        return None
    try:
        quantity = float(raw)
    except (TypeError, ValueError):
        return None
    if quantity <= 0 or quantity != quantity:
        return None
    return int(quantity) if quantity.is_integer() else round(quantity, 6)


def _analytics_kehe_case_conversion(
    quantity_ordered: Any,
    product: Optional[Dict[str, Any]],
    packaging_rows: Optional[List[Dict[str, Any]]] = None,
) -> Optional[Dict[str, Any]]:
    """Convert Analytics eaches using an explicit KeHE Product Master case pack."""
    each_quantity = _analytics_quantity(quantity_ordered)
    if each_quantity is None or not isinstance(product, dict) or not _is_kehe_storefront(product.get("storefront")):
        return None

    eaches_per_case = _analytics_quantity(product.get("case_qty"))
    if eaches_per_case is None or float(eaches_per_case) <= 1:
        return None

    eaches_per_inner_pack: Optional[float | int] = None
    inner_packs_per_case: Optional[float | int] = None
    wanted_sku = _canonical_order_sku(product.get("sku"))
    wanted_storefront = _normalize_storefront(product.get("storefront")).lower()
    for raw_row in packaging_rows or []:
        row = normalize_product_master_row(raw_row)
        if normalize_packaging_level(row.get("packaging_level")) != "Inner Pack":
            continue
        if _canonical_order_sku(row.get("sku")) != wanted_sku:
            continue
        if _normalize_storefront(row.get("storefront")).lower() != wanted_storefront:
            continue
        configured_inner_pack = _analytics_quantity(row.get("case_qty"))
        if configured_inner_pack is None or float(configured_inner_pack) <= 1:
            break
        eaches_per_inner_pack = configured_inner_pack
        calculated_inner_packs = float(eaches_per_case) / float(configured_inner_pack)
        nearest_inner_pack = round(calculated_inner_packs)
        inner_packs_per_case = (
            int(nearest_inner_pack)
            if abs(calculated_inner_packs - nearest_inner_pack) < 1e-9
            else round(calculated_inner_packs, 6)
        )
        break

    raw_case_quantity = float(each_quantity) / float(eaches_per_case)
    nearest_whole_case = round(raw_case_quantity)
    exact_case_multiple = abs(raw_case_quantity - nearest_whole_case) < 1e-9
    case_quantity = int(nearest_whole_case if exact_case_multiple else math.ceil(raw_case_quantity))
    full_cases = math.floor(raw_case_quantity)
    remainder_eaches_value = 0.0 if exact_case_multiple else float(each_quantity) - (full_cases * float(eaches_per_case))
    remainder_eaches: float | int = (
        int(round(remainder_eaches_value))
        if abs(remainder_eaches_value - round(remainder_eaches_value)) < 1e-9
        else round(remainder_eaches_value, 6)
    )

    return {
        "quantity_ordered_eaches": each_quantity,
        "quantity_ordered_cases": case_quantity,
        "quantity_ordered": case_quantity,
        "quantity_uom": "CASES",
        "eaches_per_inner_pack": eaches_per_inner_pack,
        "inner_packs_per_case": inner_packs_per_case,
        "eaches_per_case": eaches_per_case,
        "case_pack_source": "product_master",
        "case_conversion_exact": exact_case_multiple,
        "case_conversion_remainder_eaches": remainder_eaches,
    }


@app.post("/api/mpl/orders/lookup")
def lookup_mpl_order(request: Request, payload: Dict[str, Any]) -> JSONResponse:
    _require_permission(request, "generate")
    sales_order_number = str(payload.get("sales_order_number") or "").strip() if isinstance(payload, dict) else ""
    requested_ecomdash_id = str(payload.get("ecomdash_id") or "").strip() if isinstance(payload, dict) else ""
    if not sales_order_number:
        raise HTTPException(status_code=400, detail="Sales Order Number is required.")
    if len(sales_order_number) > 120 or any(ord(char) < 32 for char in sales_order_number):
        raise HTTPException(status_code=400, detail="Sales Order Number is invalid.")

    analytics_rows = _analytics_export_order_rows(request, sales_order_number)
    if not analytics_rows:
        raise HTTPException(
            status_code=404,
            detail=f"No rows were found for Sales Order Number '{sales_order_number}'.",
        )

    order_instances = _analytics_order_instance_groups(analytics_rows)
    selected_ecomdash_id = ""
    if requested_ecomdash_id:
        wanted_ecomdash_id = _canonical_order_number(requested_ecomdash_id)
        selected_instance = next(
            (
                instance
                for instance in order_instances
                if _canonical_order_number(instance.get("ecomdash_id")) == wanted_ecomdash_id
            ),
            None,
        )
        if selected_instance is None:
            raise HTTPException(
                status_code=404,
                detail=(
                    f"Ecomdash ID '{requested_ecomdash_id}' was not found for "
                    f"Sales Order Number '{sales_order_number}'."
                ),
            )
        analytics_rows = selected_instance["rows"]
        selected_ecomdash_id = str(selected_instance.get("ecomdash_id") or "")
    elif len(order_instances) > 1:
        local_file_source = ANALYTICS_ORDER_SOURCE in ANALYTICS_LOCAL_SOURCE_VALUES
        return JSONResponse(content={
            "sales_order_number": sales_order_number,
            "requires_order_selection": True,
            "order_instances": [
                _analytics_order_instance_summary(instance)
                for instance in order_instances
            ],
            "source": {
                "service": "local_file" if local_file_source else "zoho_analytics",
                "connection": "" if local_file_source else ANALYTICS_CONNECTION_LINK_NAME,
                "local_file": ANALYTICS_LOCAL_FILE.name if local_file_source and ANALYTICS_LOCAL_FILE else "",
                "workspace_id": ANALYTICS_WORKSPACE_ID,
                "view_id": ANALYTICS_VIEW_ID,
                "view_name": ANALYTICS_VIEW_NAME,
            },
        })
    elif order_instances:
        analytics_rows = order_instances[0]["rows"]
        selected_ecomdash_id = str(order_instances[0].get("ecomdash_id") or "")

    aggregated: Dict[str, Dict[str, Any]] = {}
    ignored_rows = 0
    for row in analytics_rows:
        sku = _analytics_row_value(row, ANALYTICS_SKU_COLUMN)
        quantity = _analytics_quantity(_analytics_row_value(row, ANALYTICS_QUANTITY_COLUMN))
        sku_key = _canonical_order_sku(sku)
        if not sku_key or quantity is None:
            ignored_rows += 1
            continue
        if sku_key not in aggregated:
            aggregated[sku_key] = {"sku": sku, "quantity_ordered": quantity}
        else:
            total = float(aggregated[sku_key]["quantity_ordered"]) + float(quantity)
            aggregated[sku_key]["quantity_ordered"] = int(total) if total.is_integer() else round(total, 6)

    if not aggregated:
        raise HTTPException(
            status_code=422,
            detail=(
                f"Rows were found for order '{sales_order_number}', but none had both "
                f"'{ANALYTICS_SKU_COLUMN}' and a positive '{ANALYTICS_QUANTITY_COLUMN}'."
            ),
        )

    product_rows = _datastore_load_product_master(request)
    product_source = "datastore"
    if product_rows is None:
        product_rows = _shared_product_master_file_read()
        product_source = "file"
    normalized_product_rows = _dedupe_product_master_rows(product_rows)
    eligible_products = [
        row
        for row in normalized_product_rows
        if normalize_packaging_level(row.get("packaging_level")) == "Case" and bool(row.get("in_packing_list"))
    ]
    products_by_sku: Dict[str, List[Dict[str, Any]]] = {}
    for product in eligible_products:
        key = _canonical_order_sku(product.get("sku"))
        if key:
            products_by_sku.setdefault(key, []).append(product)

    items: List[Dict[str, Any]] = []
    matched_count = 0
    ambiguous_count = 0
    converted_to_cases = 0
    partial_case_items = 0
    for sku_key, order_item in aggregated.items():
        candidates = products_by_sku.get(sku_key, [])
        if len(candidates) == 1:
            matched_count += 1
            match_status = "matched"
            product = candidates[0]
        elif len(candidates) > 1:
            ambiguous_count += 1
            match_status = "ambiguous"
            product = None
        else:
            match_status = "unmatched"
            product = None
        converted_order_item = dict(order_item)
        if product is not None and _is_kehe_storefront(product.get("storefront")):
            configured_case_pack = _analytics_quantity(product.get("case_qty"))
            if configured_case_pack is None or float(configured_case_pack) <= 1:
                raise HTTPException(
                    status_code=422,
                    detail=(
                        f"Product Master Case row for SKU '{order_item.get('sku')}' requires "
                        "Eaches / Package greater than 1 before Analytics eaches can be "
                        "converted to cases."
                    ),
                )
        conversion = _analytics_kehe_case_conversion(
            order_item.get("quantity_ordered"),
            product,
            normalized_product_rows,
        )
        if conversion:
            converted_order_item.update(conversion)
            converted_to_cases += 1
            if not conversion.get("case_conversion_exact"):
                partial_case_items += 1
        items.append({
            **converted_order_item,
            "match_status": match_status,
            "product": product,
            "candidate_storefronts": sorted({
                str(candidate.get("storefront") or "").strip()
                for candidate in candidates
                if str(candidate.get("storefront") or "").strip()
            }),
        })

    order_details = {
        field_name: next(
            (
                value
                for row in analytics_rows
                if (value := _analytics_row_value(row, column_name))
            ),
            "",
        )
        for field_name, column_name in ANALYTICS_ORDER_DETAIL_COLUMNS.items()
    }

    local_file_source = ANALYTICS_ORDER_SOURCE in ANALYTICS_LOCAL_SOURCE_VALUES
    return JSONResponse(content={
        "sales_order_number": sales_order_number,
        "order_details": order_details,
        "source": {
            "service": "local_file" if local_file_source else "zoho_analytics",
            "connection": "" if local_file_source else ANALYTICS_CONNECTION_LINK_NAME,
            "local_file": ANALYTICS_LOCAL_FILE.name if local_file_source and ANALYTICS_LOCAL_FILE else "",
            "workspace_id": ANALYTICS_WORKSPACE_ID,
            "view_id": ANALYTICS_VIEW_ID,
            "view_name": ANALYTICS_VIEW_NAME,
            "ecomdash_id": selected_ecomdash_id,
            "product_master": product_source,
        },
        "summary": {
            "analytics_rows": len(analytics_rows),
            "line_items": len(items),
            "matched_products": matched_count,
            "unmatched_products": len(items) - matched_count - ambiguous_count,
            "ambiguous_products": ambiguous_count,
            "ignored_rows": ignored_rows,
            "converted_to_cases": converted_to_cases,
            "partial_case_items": partial_case_items,
        },
        "items": items,
    })


def _read_spreadsheet_bytes(filename: str, data: bytes) -> List[Dict[str, Any]]:
    suffix = Path(filename or "").suffix.lower()
    rows: List[List[Any]] = []

    if suffix == ".csv":
        text = data.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        rows = [row for row in reader]
    elif suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook  # type: ignore
        except Exception as exc:
            raise HTTPException(status_code=500, detail="Excel upload requires openpyxl in the runtime.") from exc

        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        sheet = workbook.active
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        workbook.close()
    else:
        raise HTTPException(status_code=400, detail="Upload an .xlsx, .xlsm, or .csv file.")

    while rows and not any(str(cell or "").strip() for cell in rows[0]):
        rows.pop(0)
    if not rows:
        return []

    headers = [str(cell or "").strip() for cell in rows[0]]
    out: List[Dict[str, Any]] = []
    for raw_row in rows[1:]:
        row = {
            headers[index]: raw_row[index]
            for index in range(min(len(headers), len(raw_row)))
            if headers[index]
        }
        if any(str(value or "").strip() for value in row.values()):
            out.append(row)
    return out


def _canonical_import_key(header: str, table: str) -> str:
    key = re.sub(r"[^a-z0-9]+", "_", str(header or "").strip().lower()).strip("_")
    product_aliases = {
        "storefront": "storefront",
        "store_front": "storefront",
        "store": "storefront",
        "in_packing_list": "in_packing_list",
        "in_packing_list_": "in_packing_list",
        "packing_list": "in_packing_list",
        "include_in_packing_list": "in_packing_list",
        "include_mpl": "in_packing_list",
        "label_required": "label_required",
        "label": "label_required",
        "gtin": "gtin",
        "case_upc": "gtin",
        "upc": "gtin",
        "description": "description",
        "item_description": "description",
        "packaging_level": "packaging_level",
        "packging_level": "packaging_level",
        "level": "packaging_level",
        "l_x_w_x_h_in": "dimensions_in",
        "dimensions": "dimensions_in",
        "dimensions_in": "dimensions_in",
        "lwh_in": "dimensions_in",
        "weight_lbs": "weight_lbs",
        "weight": "weight_lbs",
        "weight_lbs_": "weight_lbs",
        "case_qty": "case_qty",
        "case_quantity": "case_qty",
        "units_per_case": "case_qty",
        "eaches_package": "case_qty",
        "eaches_per_package": "case_qty",
        "eaches_case": "case_qty",
        "eaches_per_case": "case_qty",
        "eaches_inner_pack": "case_qty",
        "eaches_per_inner_pack": "case_qty",
        "labels_unit": "labels_per_unit",
        "labels_per_unit": "labels_per_unit",
        "labels_to_print_per_unit": "labels_per_unit",
        "sku": "sku",
        "item_number": "sku",
    }
    directory_aliases = {
        "storefront": "storefront",
        "store_front": "storefront",
        "store": "storefront",
        "dc": "dc",
        "code": "dc",
        "name": "name",
        "dc_name": "name",
        "ship_from": "ship_from",
        "ship_from_address": "ship_from",
        "delivery_address": "delivery_address",
        "ship_to": "delivery_address",
        "ship_to_address": "delivery_address",
        "billing_address": "billing_address",
        "bill_to": "billing_address",
        "bill_to_address": "billing_address",
        "match_values": "match_values",
        "matching_values": "match_values",
        "gln": "match_values",
    }
    aliases = {
        "kehe_product_master": product_aliases,
        "mpl_product_master": product_aliases,
        "kehe_dc_directory": directory_aliases,
        "mpl_directory": directory_aliases,
    }
    return aliases.get(table, {}).get(key, key)


def _canonicalize_import_rows(rows: List[Dict[str, Any]], table: str) -> List[Dict[str, Any]]:
    canonical_rows: List[Dict[str, Any]] = []
    for row in rows:
        next_row: Dict[str, Any] = {}
        for key, value in row.items():
            next_row[_canonical_import_key(key, table)] = value
        canonical_rows.append(next_row)
    if table == "kehe_product_master":
        return _kehe_product_master_rows(canonical_rows)
    if table == "mpl_product_master":
        return _dedupe_product_master_rows(canonical_rows)
    if table == "kehe_dc_directory":
        return _kehe_dc_directory_rows(canonical_rows)
    if table == "mpl_directory":
        return _dedupe_dc_directory_rows(canonical_rows)
    return canonical_rows


def _preview_row_changes(
    *,
    table: str,
    current_rows: List[Dict[str, Any]],
    imported_rows: List[Dict[str, Any]],
    key_fn: Any,
) -> Dict[str, Any]:
    current_by_key = {key_fn(row): _normalized_for_audit(table, row) for row in current_rows}
    changes: List[Dict[str, Any]] = []
    added = updated = unchanged = 0

    for raw_row in imported_rows:
        key = key_fn(raw_row)
        imported = _normalized_for_audit(table, raw_row)
        current = current_by_key.get(key)
        if current is None:
            added += 1
            changes.append({
                "action": "add",
                "record_key": key,
                "record_label": _row_label(table, raw_row),
                "field": "__row__",
                "old_value": "",
                "new_value": _audit_value(imported),
            })
            continue

        row_changed = False
        for field in sorted(set(current) | set(imported)):
            old_value = _audit_value(current.get(field))
            new_value = _audit_value(imported.get(field))
            if old_value == new_value:
                continue
            row_changed = True
            changes.append({
                "action": "update",
                "record_key": key,
                "record_label": _row_label(table, raw_row),
                "field": field,
                "old_value": old_value,
                "new_value": new_value,
            })
        if row_changed:
            updated += 1
        else:
            unchanged += 1

    return {
        "summary": {
            "imported_rows": len(imported_rows),
            "added_rows": added,
            "updated_rows": updated,
            "unchanged_rows": unchanged,
            "change_count": len(changes),
        },
        "changes": changes,
    }


def _merge_import_rows(current_rows: List[Dict[str, Any]], imported_rows: List[Dict[str, Any]], key_fn: Any) -> List[Dict[str, Any]]:
    imported_by_key = {key_fn(row): row for row in imported_rows}
    merged: List[Dict[str, Any]] = []
    seen: set[str] = set()
    for row in current_rows:
        key = key_fn(row)
        if key in imported_by_key:
            merged.append(imported_by_key[key])
            seen.add(key)
        else:
            merged.append(row)
    for key, row in imported_by_key.items():
        if key not in seen:
            merged.append(row)
    return merged


async def _preview_excel_import(request: Request, upload: UploadFile, table: str) -> JSONResponse:
    _require_permission(request, "table_crud")
    data = await upload.read()
    await upload.close()
    raw_rows = _read_spreadsheet_bytes(upload.filename or "upload.xlsx", data)
    imported_rows = _canonicalize_import_rows(raw_rows, table)
    if table in {"kehe_product_master", "mpl_product_master"}:
        current_rows = _datastore_load_product_master(request)
        if current_rows is None:
            current_rows = _shared_product_master_file_read()
        if table == "kehe_product_master":
            current_rows = _kehe_product_master_rows(current_rows)
        else:
            current_rows = _dedupe_product_master_rows(current_rows)
        key_fn = _product_row_key
    else:
        current_rows = _datastore_load_dc_directory(request)
        if current_rows is None:
            current_rows = _shared_dc_directory_file_read()
        if table == "kehe_dc_directory":
            current_rows = _kehe_dc_directory_rows(current_rows)
        else:
            current_rows = _dedupe_dc_directory_rows(current_rows)
        key_fn = _dc_row_key

    preview = _preview_row_changes(
        table=table,
        current_rows=current_rows,
        imported_rows=imported_rows,
        key_fn=key_fn,
    )
    batch_id = uuid.uuid4().hex
    return JSONResponse(content={
        "batch_id": batch_id,
        "filename": upload.filename,
        "table": table,
        "rows": imported_rows,
        **preview,
    })


@app.post("/api/mpl/product-master/import-preview")
async def preview_mpl_product_master_import(request: Request, file: UploadFile = File(...)) -> JSONResponse:
    return await _preview_excel_import(request, file, "mpl_product_master")


@app.post("/api/mpl/directory/import-preview")
async def preview_mpl_directory_import(request: Request, file: UploadFile = File(...)) -> JSONResponse:
    return await _preview_excel_import(request, file, "mpl_directory")


@app.post("/api/mpl/product-master/import-confirm")
async def confirm_mpl_product_master_import(request: Request, payload: Dict[str, Any]) -> JSONResponse:
    _require_permission(request, "table_crud")
    imported_rows = payload.get("rows") if isinstance(payload, dict) else []
    if not isinstance(imported_rows, list):
        imported_rows = []
    imported_rows = _canonicalize_import_rows([r for r in imported_rows if isinstance(r, dict)], "mpl_product_master")
    current_rows = _datastore_load_product_master(request)
    if current_rows is None:
        current_rows = _shared_product_master_file_read()
    merged_rows = _merge_import_rows(current_rows, imported_rows, _product_row_key)
    return await save_mpl_product_master(request, {
        "rows": merged_rows,
        "source": "excel_import",
        "batch_id": str(payload.get("batch_id") or uuid.uuid4().hex),
        "filename": str(payload.get("filename") or ""),
    })


@app.post("/api/mpl/directory/import-confirm")
async def confirm_mpl_directory_import(request: Request, payload: Dict[str, Any]) -> JSONResponse:
    _require_permission(request, "table_crud")
    imported_rows = payload.get("rows") if isinstance(payload, dict) else []
    if not isinstance(imported_rows, list):
        imported_rows = []
    imported_rows = _canonicalize_import_rows([r for r in imported_rows if isinstance(r, dict)], "mpl_directory")
    current_rows = _datastore_load_dc_directory(request)
    if current_rows is None:
        current_rows = _shared_dc_directory_file_read()
    merged_rows = _merge_import_rows(current_rows, imported_rows, _dc_row_key)
    return await save_mpl_directory(request, {
        "rows": merged_rows,
        "source": "excel_import",
        "batch_id": str(payload.get("batch_id") or uuid.uuid4().hex),
        "filename": str(payload.get("filename") or ""),
    })


def _mpl_drafts_datastore_table(request: Optional[Request]) -> Any:
    if request is None:
        if _store_requires_datastore(MPL_DRAFTS_STORE):
            _raise_datastore_unavailable(MPL_DRAFTS_TABLE, "connect to it")
        return None
    return _datastore_table_named(
        request,
        MPL_DRAFTS_TABLE,
        MPL_DRAFTS_STORE,
    )


def _datastore_row_to_mpl_draft(row: Dict[str, Any]) -> Dict[str, Any]:
    draft_raw = row.get("DRAFT_JSON") or row.get("draft") or {}
    draft = draft_raw if isinstance(draft_raw, dict) else {}
    if isinstance(draft_raw, str) and draft_raw.strip():
        try:
            encoded = draft_raw.strip()
            if encoded.startswith("zlib:"):
                compressed = base64.urlsafe_b64decode(encoded[5:].encode("ascii"))
                encoded = zlib.decompress(compressed).decode("utf-8")
            parsed = json.loads(encoded)
            if isinstance(parsed, dict):
                draft = parsed
        except Exception:
            draft = {}
    return {
        "id": row.get("DRAFT_ID") or row.get("id") or row.get("ROWID") or "",
        "name": row.get("NAME") or row.get("name") or "",
        "created_at": row.get("CREATED_AT") or row.get("created_at") or "",
        "updated_at": row.get("UPDATED_AT") or row.get("updated_at") or "",
        "created_by": row.get("CREATED_BY") or row.get("created_by") or "",
        "updated_by": row.get("UPDATED_BY") or row.get("updated_by") or "",
        "draft": draft,
    }


def _mpl_draft_to_datastore_row(record: Dict[str, Any]) -> Dict[str, Any]:
    draft_json = json.dumps(record.get("draft") or {}, sort_keys=True, separators=(",", ":"))
    compressed_draft = base64.urlsafe_b64encode(
        zlib.compress(draft_json.encode("utf-8"), level=9)
    ).decode("ascii")
    return {
        "DRAFT_ID": str(record.get("id") or uuid.uuid4().hex),
        "NAME": str(record.get("name") or ""),
        "CREATED_AT": str(record.get("created_at") or _now_iso()),
        "UPDATED_AT": str(record.get("updated_at") or _now_iso()),
        "DRAFT_JSON": f"zlib:{compressed_draft}",
        "IS_ACTIVE": True,
    }


def _mpl_draft_for_storage(draft: Dict[str, Any]) -> Dict[str, Any]:
    """Remove regenerable data that can overflow the Data Store JSON text column."""
    def clean(value: Any, key: str = "") -> Any:
        normalized_key = str(key or "").strip().lower()
        if normalized_key == "product_master":
            return None
        if normalized_key in {"_tihi_snapshot", "sheet_image_data_url", "image_data_url"}:
            return None
        if isinstance(value, dict):
            return {
                child_key: cleaned
                for child_key, child_value in value.items()
                if (cleaned := clean(child_value, child_key)) is not None
            }
        if isinstance(value, list):
            return [cleaned for item in value if (cleaned := clean(item)) is not None]
        return value

    cleaned_draft = clean(draft)
    return cleaned_draft if isinstance(cleaned_draft, dict) else {}


def _datastore_load_mpl_drafts(request: Optional[Request]) -> Optional[List[Dict[str, Any]]]:
    table_service = _mpl_drafts_datastore_table(request)
    if table_service is None:
        return None
    try:
        raw_rows = _datastore_get_raw_rows(table_service)
        active_rows = [
            row for row in raw_rows
            if str(row.get("IS_ACTIVE", True)).lower() not in {"false", "0", "no"}
        ]
        return [_datastore_row_to_mpl_draft(row) for row in active_rows]
    except Exception as exc:
        if _store_requires_datastore(MPL_DRAFTS_STORE):
            _raise_datastore_unavailable(MPL_DRAFTS_TABLE, "read from it", exc)
        return None


def _datastore_save_mpl_drafts(request: Optional[Request], drafts: List[Dict[str, Any]]) -> bool:
    table_service = _mpl_drafts_datastore_table(request)
    if table_service is None:
        return False
    try:
        existing_rows = _datastore_get_raw_rows(table_service)
        row_ids = [row.get("ROWID") for row in existing_rows if row.get("ROWID")]
        for batch in _chunked(row_ids, 200):
            table_service.delete_rows(batch)
        insert_rows = [_mpl_draft_to_datastore_row(record) for record in drafts]
        for batch in _chunked(insert_rows, 100):
            if batch:
                table_service.insert_rows(batch)
        return True
    except Exception as exc:
        if _store_requires_datastore(MPL_DRAFTS_STORE):
            _raise_datastore_unavailable(MPL_DRAFTS_TABLE, "write to it", exc)
        return False


def _mpl_drafts_read(request: Optional[Request] = None) -> List[Dict[str, Any]]:
    datastore_rows = _datastore_load_mpl_drafts(request)
    if datastore_rows is not None:
        return datastore_rows
    try:
        if not MPL_DRAFTS_FILE.exists():
            return []
        data = json.loads(MPL_DRAFTS_FILE.read_text(encoding="utf-8"))
        drafts = data.get("drafts") if isinstance(data, dict) else data
        if not isinstance(drafts, list):
            return []
        return [draft for draft in drafts if isinstance(draft, dict)]
    except Exception:
        return []


def _mpl_drafts_write(drafts: List[Dict[str, Any]], request: Optional[Request] = None) -> List[Dict[str, Any]]:
    if _datastore_save_mpl_drafts(request, drafts):
        return drafts
    MPL_DRAFTS_FILE.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "drafts": drafts,
        "updated_at": _now_iso(),
    }
    MPL_DRAFTS_FILE.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    return drafts


def _mpl_draft_summary(record: Dict[str, Any]) -> Dict[str, Any]:
    draft = record.get("draft") if isinstance(record.get("draft"), dict) else {}
    mpl = {}
    packing_lists = draft.get("packing_lists") if isinstance(draft, dict) else []
    if isinstance(packing_lists, list) and packing_lists:
        mpl = packing_lists[0] if isinstance(packing_lists[0], dict) else {}
    items = mpl.get("items") if isinstance(mpl.get("items"), list) else []
    created_by = (
        record.get("created_by")
        or draft.get("_saved_draft_created_by", "")
        or draft.get("_saved_by", "")
    )
    updated_by = (
        record.get("updated_by")
        or draft.get("_saved_draft_updated_by", "")
        or draft.get("_saved_by", "")
    )
    return {
        "id": record.get("id", ""),
        "name": record.get("name", ""),
        "created_at": record.get("created_at", ""),
        "updated_at": record.get("updated_at", ""),
        "created_by": created_by,
        "updated_by": updated_by,
        "customer_po_number": mpl.get("customer_po_number", ""),
        "ship_to": str(mpl.get("ship_to", "")).split("\n")[0] if mpl.get("ship_to") else "",
        "total_pallets": mpl.get("total_pallets", ""),
        "item_count": len(items),
    }


@app.get("/api/kehe/mpl-drafts")
async def list_kehe_mpl_drafts(request: Request) -> JSONResponse:
    _require_permission(request, "view")
    drafts = _mpl_drafts_read(request)
    summaries = [_mpl_draft_summary(record) for record in drafts]
    summaries.sort(key=lambda row: str(row.get("updated_at", "")), reverse=True)
    return JSONResponse(content={"drafts": summaries})


@app.get("/api/kehe/mpl-drafts/{draft_id}")
async def get_kehe_mpl_draft(request: Request, draft_id: str) -> JSONResponse:
    _require_permission(request, "view")
    for record in _mpl_drafts_read(request):
        if str(record.get("id")) == draft_id:
            return JSONResponse(content={"draft": record})
    raise HTTPException(status_code=404, detail="Saved MPL draft not found.")


@app.post("/api/kehe/mpl-drafts")
async def save_kehe_mpl_draft(request: Request, payload: Dict[str, Any]) -> JSONResponse:
    _require_permission(request, "save_mpl")
    draft = payload.get("draft") if isinstance(payload, dict) else None
    if not isinstance(draft, dict):
        raise HTTPException(status_code=400, detail="MPL draft payload is required.")

    storage_draft = _mpl_draft_for_storage(draft)
    drafts = _mpl_drafts_read(request)
    draft_id = str(payload.get("id") or draft.get("_saved_draft_id") or uuid.uuid4().hex)
    now = _now_iso()
    name = str(payload.get("name") or draft.get("_saved_draft_name") or "").strip()
    if not name:
        packing_lists = draft.get("packing_lists") if isinstance(draft.get("packing_lists"), list) else []
        first = packing_lists[0] if packing_lists and isinstance(packing_lists[0], dict) else {}
        name = str(first.get("customer_po_number") or first.get("id") or "Untitled MPL").strip()

    old_record = next((record for record in drafts if str(record.get("id")) == draft_id), None)
    created_at = old_record.get("created_at") if old_record else now
    actor = _request_actor(request)
    actor_label = str(actor.get("email") or actor.get("name") or "Local user")
    created_by = (
        old_record.get("created_by")
        if old_record else ""
    ) or str(draft.get("_saved_draft_created_by") or actor_label)
    storage_draft["_saved_draft_created_by"] = created_by
    storage_draft["_saved_draft_updated_by"] = actor_label
    record = {
        "id": draft_id,
        "name": name,
        "created_at": created_at,
        "updated_at": now,
        "created_by": created_by,
        "updated_by": actor_label,
        "draft": storage_draft,
    }
    next_drafts = [record if str(existing.get("id")) == draft_id else existing for existing in drafts]
    if old_record is None:
        next_drafts.append(record)
    _mpl_drafts_write(next_drafts, request)

    _audit_log_append([{
        "id": uuid.uuid4().hex,
        "timestamp": now,
        "actor": actor,
        "table": "kehe_mpl_drafts",
        "action": "update" if old_record else "create",
        "record_key": draft_id,
        "record_label": name,
        "field": "draft",
        "old_value": old_record.get("name", "") if old_record else "",
        "new_value": name,
        "source": "mpl_save",
        "batch_id": "",
        "filename": "",
    }], request=request)

    return JSONResponse(content={"draft": record, "saved": True})


@app.post("/api/kehe/mpl-drafts/{draft_id}/delete")
@app.delete("/api/kehe/mpl-drafts/{draft_id}")
async def delete_kehe_mpl_draft(request: Request, draft_id: str) -> JSONResponse:
    _require_permission(request, "delete_mpl")
    drafts = _mpl_drafts_read(request)
    old_record = next((record for record in drafts if str(record.get("id")) == draft_id), None)
    if old_record is None:
        raise HTTPException(status_code=404, detail="Saved MPL draft not found.")

    next_drafts = [record for record in drafts if str(record.get("id")) != draft_id]
    _mpl_drafts_write(next_drafts, request)

    now = _now_iso()
    _audit_log_append([{
        "id": uuid.uuid4().hex,
        "timestamp": now,
        "actor": _request_actor(request),
        "table": "kehe_mpl_drafts",
        "action": "delete",
        "record_key": draft_id,
        "record_label": old_record.get("name", ""),
        "field": "draft",
        "old_value": old_record.get("name", ""),
        "new_value": "",
        "source": "mpl_delete",
        "batch_id": "",
        "filename": "",
    }], request=request)

    return JSONResponse(content={"deleted": True, "id": draft_id})


# ---------------------------------------------------------------------------
# BACKEND SECTION 5: shared result endpoints for frontend polling/download.
# ---------------------------------------------------------------------------
@app.get("/results/{result_id}/status")
def get_result_status(request: Request, result_id: str) -> JSONResponse:
    _require_permission(request, "view")
    job = RESULT_JOBS.get(result_id)
    if job is None:
        raise HTTPException(status_code=404, detail="Generation result not found.")

    return JSONResponse(
        content={
            "result_id": result_id,
            "kit": job.get("kit"),
            "status": job.get("status", "processing"),
            "detail": job.get("detail", "Processing…"),
            "report": job.get("report"),
            "file_ready": bool(job.get("output_path")),
        }
    )


@app.get("/results/{result_id}/report")
def get_result_report(request: Request, result_id: str) -> JSONResponse:
    _require_permission(request, "view")
    report = RESULT_REPORTS.get(result_id)
    if report is None:
        job = RESULT_JOBS.get(result_id)
        if job is not None:
            report = job.get("report")
    if report is None:
        raise HTTPException(status_code=404, detail="Match report not found.")
    return JSONResponse(content=report)


@app.get("/results/{result_id}/file")
def get_result_file(request: Request, result_id: str) -> FileResponse:
    _require_permission(request, "view")
    job = RESULT_JOBS.get(result_id)
    if job is None:
        raise HTTPException(status_code=404, detail="Generated file not found.")
    if job.get("status") != "complete" or not job.get("output_path"):
        raise HTTPException(status_code=409, detail="PDF is not ready yet.")

    kit = str(job.get("kit") or "michaels")
    filename = (
        job.get("output_filename")
        or KIT_CONFIG.get(kit, KIT_CONFIG["michaels"])["output_filename"]
    )
    return FileResponse(
        path=job["output_path"],
        media_type="application/pdf",
        filename=filename,
        headers={"Access-Control-Expose-Headers": "X-Result-Id"},
    )


# ---------------------------------------------------------------------------
# BACKEND SECTION 6: kit-specific generation endpoints.
# Frontend buttons choose one of these endpoints.
# ---------------------------------------------------------------------------
@app.post("/generate/{kit}")
async def generate_for_kit(
    kit: str,
    request: Request,
    xml_files: List[UploadFile] = File(...),
    pdf_files: Optional[List[UploadFile]] = File(default=None),
    mode: Optional[str] = Form(default="xml"),
) -> JSONResponse:
    _require_permission(request, "generate")
    kit = normalize_kit(kit)
    if kit not in KIT_CONFIG:
        raise HTTPException(status_code=404, detail="Unknown label kit. Use 'michaels' or 'kehe'.")
    if not xml_files:
        raise HTTPException(status_code=400, detail="At least one XML file is required.")
    if kit == "michaels" and not pdf_files:
        raise HTTPException(status_code=400, detail="Michaels requires at least one shipping-label PDF file.")

    # mode is accepted for compatibility with older frontend bundles.
    _ = mode

    temp_dir = Path(tempfile.mkdtemp(prefix=KIT_CONFIG[kit]["temp_prefix"]))
    try:
        xml_paths: List[str] = []
        pdf_paths: List[str] = []

        for upload in xml_files:
            if not (upload.filename or "").lower().endswith(".xml"):
                raise HTTPException(status_code=400, detail=f"Invalid XML file: {upload.filename}")
            out_path = temp_dir / sanitize_filename(upload.filename or "input.xml")
            await save_upload_file(upload, out_path)
            xml_paths.append(str(out_path))

        if pdf_files:
            for upload in pdf_files:
                if not (upload.filename or "").lower().endswith(".pdf"):
                    raise HTTPException(status_code=400, detail=f"Invalid PDF file: {upload.filename}")
                out_path = temp_dir / sanitize_filename(upload.filename or "label.pdf")
                await save_upload_file(upload, out_path)
                pdf_paths.append(str(out_path))

        result_id = create_result_job(temp_dir, kit)
        if kit == "michaels":
            worker = threading.Thread(
                target=run_michaels_generation_job,
                args=(result_id, xml_paths, pdf_paths),
                daemon=True,
            )
        else:
            worker = threading.Thread(
                target=run_kehe_generation_job,
                args=(result_id, xml_paths),
                daemon=True,
            )
        worker.start()

        return JSONResponse(
            content={
                "result_id": result_id,
                "kit": kit,
                "status": "processing",
                "detail": f"{KIT_CONFIG[kit]['label']} generation started…",
            },
            headers={
                "X-Result-Id": result_id,
                "Access-Control-Expose-Headers": "X-Result-Id",
            },
        )
    except HTTPException:
        shutil.rmtree(temp_dir, ignore_errors=True)
        raise
    except Exception as exc:
        shutil.rmtree(temp_dir, ignore_errors=True)
        return JSONResponse(status_code=500, content={"detail": str(exc)})


# ---------------------------------------------------------------------------
# BACKEND SECTION 6B: KeHE document prepare endpoints (draft JSON only).
# These parse XML and return editable JSON. They do NOT generate PDFs.
# ---------------------------------------------------------------------------
@app.post("/prepare/kehe/pallet-label")
async def prepare_kehe_pallet_label(
    request: Request,
    xml_files: List[UploadFile] = File(...),
) -> JSONResponse:
    _require_permission(request, "generate")
    if not xml_files:
        raise HTTPException(status_code=400, detail="At least one XML file is required.")
    temp_dir = Path(tempfile.mkdtemp(prefix="kehe_pallet_prepare_"))
    try:
        xml_paths: List[str] = []
        for upload in xml_files:
            if not (upload.filename or "").lower().endswith(".xml"):
                raise HTTPException(status_code=400, detail=f"Invalid XML file: {upload.filename}")
            out_path = temp_dir / sanitize_filename(upload.filename or "input.xml")
            await save_upload_file(upload, out_path)
            xml_paths.append(str(out_path))
        _sync_kehe_dc_directory_for_pipeline(request)
        draft = build_kehe_pallet_label_draft(xml_paths)
        # Attach extracted_headers for the frontend Extracted Data table
        draft["extracted_headers"] = [
            {
                "source_file":             p.get("id", ""),
                "customer_po_numbers":     p.get("customer_po_numbers", ""),
                "pro_number":              p.get("pro_number", ""),
                "bol_number":              p.get("bol_number", ""),
                "ship_date":               p.get("date", ""),
                "expected_delivery_date":  p.get("expected_delivery_date", ""),
                "carrier":                 p.get("carrier", ""),
                "total_weight":            "",
                "carton_count":            p.get("carton_count", ""),
                "total_pallets":           p.get("total_pallets", ""),
                "ship_via":                p.get("carrier", ""),
                "dc":                      p.get("dc", ""),
                "ship_to_name":            (p.get("ship_to") or "").split("\n")[0],
            }
            for p in (draft.get("pallets") or [])
        ]
        return JSONResponse(content=draft)
    except HTTPException:
        raise
    except Exception as exc:
        print(f"DEBUG prepare_kehe_pallet_label error: {exc}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error preparing pallet label draft: {str(exc)}")
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)


@app.post("/prepare/kehe/master-packing-list")
async def prepare_kehe_master_packing_list(
    request: Request,
    xml_files: List[UploadFile] = File(...),
    product_master_json: Optional[str] = Form(default="[]"),
) -> JSONResponse:
    _require_permission(request, "generate")
    if not xml_files:
        raise HTTPException(status_code=400, detail="At least one XML file is required.")
    temp_dir = Path(tempfile.mkdtemp(prefix="kehe_mpl_prepare_"))
    try:
        xml_paths: List[str] = []
        for upload in xml_files:
            if not (upload.filename or "").lower().endswith(".xml"):
                raise HTTPException(status_code=400, detail=f"Invalid XML file: {upload.filename}")
            out_path = temp_dir / sanitize_filename(upload.filename or "input.xml")
            await save_upload_file(upload, out_path)
            xml_paths.append(str(out_path))
        product_master_rows = parse_product_master_json(product_master_json)
        if not product_master_rows:
            product_master_rows = _datastore_load_product_master(request)
            if product_master_rows is None:
                product_master_rows = _shared_product_master_file_read()
            product_master_rows = _kehe_product_master_rows(product_master_rows)
        else:
            product_master_rows = _kehe_product_master_rows(product_master_rows)
        _sync_kehe_dc_directory_for_pipeline(request)
        draft = build_kehe_master_packing_list_draft(xml_paths, product_master_rows=product_master_rows)
        # Attach extracted_headers for the frontend Extracted Data table
        draft["extracted_headers"] = [
            {
                "source_file":             (m.get("source_files") or [""])[0],
                "customer_po_numbers":     m.get("customer_po_number", ""),
                "pro_number":              m.get("pro_number", ""),
                "bol_number":              "",
                "ship_date":               m.get("est_ship_date", ""),
                "expected_delivery_date":  m.get("expected_delivery_date", ""),
                "carrier":                 m.get("ship_via", ""),
                "total_weight":            m.get("total_weight", ""),
                "carton_count":            str(len(m.get("items") or [])),
                "total_pallets":           m.get("total_pallets", ""),
                "ship_via":                m.get("ship_via", ""),
                "dc":                      m.get("dc", ""),
                "ship_to_name":            (m.get("ship_to") or "").split("\n")[0],
            }
            for m in (draft.get("packing_lists") or [])
        ]
        return JSONResponse(content=draft)
    except HTTPException:
        raise
    except Exception as exc:
        print(f"DEBUG prepare_kehe_master_packing_list error: {exc}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error preparing master packing list draft: {str(exc)}")
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)


@app.post("/prepare/kehe/pack-labels")
async def prepare_kehe_pack_labels(
    request: Request,
    xml_files: List[UploadFile] = File(...),
    product_master_json: Optional[str] = Form(default="[]"),
) -> JSONResponse:
    _require_permission(request, "generate")
    if not xml_files:
        raise HTTPException(status_code=400, detail="At least one XML file is required.")
    temp_dir = Path(tempfile.mkdtemp(prefix="kehe_pack_labels_prepare_"))
    try:
        xml_paths: List[str] = []
        for upload in xml_files:
            if not (upload.filename or "").lower().endswith(".xml"):
                raise HTTPException(status_code=400, detail=f"Invalid XML file: {upload.filename}")
            out_path = temp_dir / sanitize_filename(upload.filename or "input.xml")
            await save_upload_file(upload, out_path)
            xml_paths.append(str(out_path))

        product_master_rows = parse_product_master_json(product_master_json)
        if not product_master_rows:
            product_master_rows = _datastore_load_product_master(request)
            if product_master_rows is None:
                product_master_rows = _shared_product_master_file_read()
            product_master_rows = _kehe_product_master_rows(product_master_rows)
        else:
            product_master_rows = _kehe_product_master_rows(product_master_rows)
        _sync_kehe_dc_directory_for_pipeline(request)
        draft = build_kehe_pack_label_draft(xml_paths, product_master_rows=product_master_rows)
        return JSONResponse(content=draft)
    except HTTPException:
        raise
    except Exception as exc:
        print(f"DEBUG prepare_kehe_pack_labels error: {exc}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error preparing pack label draft: {str(exc)}")
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)


# ---------------------------------------------------------------------------
# BACKEND SECTION 6C: KeHE document render workers.
# ---------------------------------------------------------------------------
def run_kehe_pallet_label_render_job(result_id: str, draft: Dict[str, Any]) -> None:
    job = RESULT_JOBS[result_id]
    temp_dir = Path(job["temp_dir"])
    output_path = temp_dir / KIT_CONFIG["kehe_pallet_label"]["output_filename"]
    try:
        def _progress(message: str) -> None:
            update_result_job(result_id, detail=message)

        update_result_job(result_id, detail="Rendering edited KeHE Pallet Label PDF\u2026")
        report = render_kehe_pallet_label_pdf(
            draft=draft,
            out_pdf=str(output_path),
            progress_callback=_progress,
        )
        if not output_path.exists():
            raise RuntimeError("Output PDF was not generated.")
        RESULT_REPORTS[result_id] = report
        update_result_job(
            result_id,
            status="complete",
            detail="KeHE Pallet Label generated successfully.",
            report=report,
            output_path=str(output_path),
        )
    except Exception as exc:
        update_result_job(result_id, status="error", detail=str(exc))


def run_kehe_master_packing_list_render_job(result_id: str, draft: Dict[str, Any]) -> None:
    job = RESULT_JOBS[result_id]
    temp_dir = Path(job["temp_dir"])
    output_path = temp_dir / KIT_CONFIG["kehe_master_packing_list"]["output_filename"]
    try:
        def _progress(message: str) -> None:
            update_result_job(result_id, detail=message)

        update_result_job(result_id, detail="Rendering edited KeHE Master Packing List PDF\u2026")
        report = render_kehe_master_packing_list_pdf(
            draft=draft,
            out_pdf=str(output_path),
            progress_callback=_progress,
        )
        if not output_path.exists():
            raise RuntimeError("Output PDF was not generated.")
        RESULT_REPORTS[result_id] = report
        update_result_job(
            result_id,
            status="complete",
            detail="KeHE Master Packing List generated successfully.",
            report=report,
            output_path=str(output_path),
        )
    except Exception as exc:
        update_result_job(result_id, status="error", detail=str(exc))


def run_kehe_pack_label_render_job(result_id: str, draft: Dict[str, Any]) -> None:
    job = RESULT_JOBS[result_id]
    temp_dir = Path(job["temp_dir"])
    output_path = temp_dir / KIT_CONFIG["kehe_pack_labels"]["output_filename"]
    try:
        def _progress(message: str) -> None:
            update_result_job(result_id, detail=message)

        update_result_job(result_id, detail="Rendering edited KeHE Pack Labels PDF…")
        report = render_kehe_pack_label_pdf(
            draft=draft,
            out_pdf=str(output_path),
            progress_callback=_progress,
        )
        if not output_path.exists():
            raise RuntimeError("Output PDF was not generated.")
        RESULT_REPORTS[result_id] = report
        update_result_job(
            result_id,
            status="complete",
            detail="KeHE Pack Labels generated successfully.",
            report=report,
            output_path=str(output_path),
        )
    except Exception as exc:
        update_result_job(result_id, status="error", detail=str(exc))


# ---------------------------------------------------------------------------
# BACKEND SECTION 6D: KeHE document render endpoints.
# ---------------------------------------------------------------------------
@app.post("/render/kehe/pallet-label")
async def render_kehe_pallet_label_endpoint(request: Request, draft: Dict[str, Any]) -> JSONResponse:
    _require_permission(request, "generate")
    temp_dir = Path(tempfile.mkdtemp(prefix=KIT_CONFIG["kehe_pallet_label"]["temp_prefix"]))
    result_id = create_result_job(
        temp_dir,
        "kehe_pallet_label",
        output_filename=KIT_CONFIG["kehe_pallet_label"]["output_filename"],
    )
    worker = threading.Thread(
        target=run_kehe_pallet_label_render_job,
        args=(result_id, draft),
        daemon=True,
    )
    worker.start()
    return JSONResponse(
        content={
            "result_id": result_id,
            "kit": "kehe_pallet_label",
            "status": "processing",
            "detail": "KeHE Pallet Label generation started\u2026",
        },
        headers={
            "X-Result-Id": result_id,
            "Access-Control-Expose-Headers": "X-Result-Id",
        },
    )


@app.post("/render/kehe/master-packing-list")
async def render_kehe_master_packing_list_endpoint(request: Request, draft: Dict[str, Any]) -> JSONResponse:
    _require_permission(request, "generate")
    temp_dir = Path(tempfile.mkdtemp(prefix=KIT_CONFIG["kehe_master_packing_list"]["temp_prefix"]))
    result_id = create_result_job(
        temp_dir,
        "kehe_master_packing_list",
        output_filename=KIT_CONFIG["kehe_master_packing_list"]["output_filename"],
    )
    worker = threading.Thread(
        target=run_kehe_master_packing_list_render_job,
        args=(result_id, draft),
        daemon=True,
    )
    worker.start()
    return JSONResponse(
        content={
            "result_id": result_id,
            "kit": "kehe_master_packing_list",
            "status": "processing",
            "detail": "KeHE Master Packing List generation started\u2026",
        },
        headers={
            "X-Result-Id": result_id,
            "Access-Control-Expose-Headers": "X-Result-Id",
        },
    )


@app.post("/render/kehe/pack-labels")
async def render_kehe_pack_labels_endpoint(request: Request, draft: Dict[str, Any]) -> JSONResponse:
    _require_permission(request, "generate")
    temp_dir = Path(tempfile.mkdtemp(prefix=KIT_CONFIG["kehe_pack_labels"]["temp_prefix"]))
    result_id = create_result_job(
        temp_dir,
        "kehe_pack_labels",
        output_filename=KIT_CONFIG["kehe_pack_labels"]["output_filename"],
    )
    worker = threading.Thread(
        target=run_kehe_pack_label_render_job,
        args=(result_id, draft),
        daemon=True,
    )
    worker.start()
    return JSONResponse(
        content={
            "result_id": result_id,
            "kit": "kehe_pack_labels",
            "status": "processing",
            "detail": "KeHE Pack Labels generation started…",
        },
        headers={
            "X-Result-Id": result_id,
            "Access-Control-Expose-Headers": "X-Result-Id",
        },
    )


# ---------------------------------------------------------------------------
# BACKEND SECTION 7: utility helpers.
# ---------------------------------------------------------------------------
async def save_upload_file(upload: UploadFile, destination: Path) -> None:
    with destination.open("wb") as f:
        while True:
            chunk = await upload.read(1024 * 1024)
            if not chunk:
                break
            f.write(chunk)
    await upload.close()


def combine_shipping_pdfs(pdf_paths: List[Path], combined_path: Path) -> Path:
    if len(pdf_paths) == 1:
        return pdf_paths[0]

    merged = fitz.open()
    try:
        for pdf_path in pdf_paths:
            src = fitz.open(pdf_path)
            try:
                merged.insert_pdf(src)
            finally:
                src.close()
        merged.save(combined_path)
    finally:
        merged.close()
    return combined_path


def sanitize_filename(name: str) -> str:
    keep = []
    for ch in name:
        if ch.isalnum() or ch in ("-", "_", ".", " "):
            keep.append(ch)
        else:
            keep.append("_")
    return "".join(keep)


def normalize_kit(value: str) -> str:
    normalized = (value or "").strip().lower().replace("_", "-")
    aliases = {
        "michael": "michaels",
        "michaels-label-kit": "michaels",
        "michaels-labelkit": "michaels",
        "michaels-dts": "michaels",
        "kehe-label-kit": "kehe",
        "kehe-labelkit": "kehe",
        "kehe-gs1": "kehe",
    }
    return aliases.get(normalized, normalized)


def serve_frontend_index() -> HTMLResponse:
    index_path = FRONTEND_DIST / "index.html"
    if not index_path.exists():
        return HTMLResponse(
            "<h1>Frontend build not found</h1>"
            "<p>The bundled frontend/dist folder is missing.</p>",
            status_code=500,
        )
    html = index_path.read_text(encoding="utf-8")
    return HTMLResponse(
        html,
        media_type="text/html",
        headers={
            "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
            "Pragma": "no-cache",
            "Expires": "0",
        },
    )


@app.get("/{full_path:path}")
async def spa_fallback(full_path: str, request: Request):
    _ = request
    if full_path.startswith(("api", "generate", "prepare", "render", "results", "health", "docs", "openapi.json", "redoc", "accounts/")):
        raise HTTPException(status_code=404, detail="Not found")

    requested_path = FRONTEND_DIST / full_path
    if requested_path.is_file():
        if requested_path.name.lower() == "index.html":
            return serve_frontend_index()
        return FileResponse(requested_path)

    return serve_frontend_index()


if __name__ == "__main__":
    import uvicorn

    uvicorn.run(
        "server:app",
        host="0.0.0.0",
        port=DEFAULT_PORT,
        reload=False,
        log_level="info",
    )


